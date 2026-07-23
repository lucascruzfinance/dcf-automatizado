"""Exportador Excel "Modelo" (Prompt 9.0.5 — padrao >= Direcional).

Gera ``outputs/excel/<TICKER>_dcf.xlsx`` para QUALQUER nao-financeira com as
8 abas: Capa, Premissas, **Modelo** (DRE pre-D&A + BP aberto + DFC indireto +
WK + Divida + PP&E, historicos CVM + 8 anos projetados), **FCFF** e **FCFE**
(abas SEPARADAS, ambas referenciando a aba Modelo como fonte), Macro,
Sensibilidades e Avisos.

Convencao de CORES de Lucas (Principio 11 — substitui a convencao WSP):
- fonte **AZUL**  = dado HISTORICO (vindo da CVM);
- fonte **VERDE** = PREMISSA que o usuario escolhe (e referencias diretas a
  premissas dentro do Modelo);
- fonte **PRETO** = resultado de FORMULA nativa do Excel.

O exportador e um CONSUMIDOR puro do motor: nenhum numero nasce aqui. Cada
celula projetada recebe uma formula nativa que reproduz o calculo do motor
(mecanismo ``escrever_calculo``); quando a formula NAO reproduz o valor
persistido (salvaguardas, pisos, truncamentos), a celula recebe o VALOR do
motor — o Excel nunca mostra um numero diferente do pipeline. Alterar uma
premissa VERDE (ex.: margem bruta do ano 3) propaga pelas formulas ate o
Target das abas FCFF e FCFE (modelo vivo).
"""

from __future__ import annotations

import logging
import math
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

RAIZ_PROJETO = Path(__file__).resolve().parents[2]
if str(RAIZ_PROJETO) not in sys.path:
    sys.path.insert(0, str(RAIZ_PROJETO))

try:
    from src.metricas.metricas_historicas import montar_series_anuais
    from src.projecao.projetor_dre import (
        ALIQUOTA_RET_RECEITA,
        HORIZONTE_PROJECAO,
        carregar_json,
        carregar_metadados,
        empresa_usa_ret,
        normalizar_ticker,
        resolver_raiz,
    )
    from src.valuation.calculador_vt import (
        carregar_convencao_desconto,
        expoente_desconto,
    )
except ModuleNotFoundError as erro:
    if erro.name != "src":
        raise
    from src.metricas.metricas_historicas import montar_series_anuais
    from src.projecao.projetor_dre import (
        ALIQUOTA_RET_RECEITA,
        HORIZONTE_PROJECAO,
        carregar_json,
        carregar_metadados,
        empresa_usa_ret,
        normalizar_ticker,
        resolver_raiz,
    )
    from src.valuation.calculador_vt import (
        carregar_convencao_desconto,
        expoente_desconto,
    )

logger = logging.getLogger(__name__)

# Nomes e ordem das 8 abas (Prompt 9.0.5). Fonte unica para o .xlsx e o
# preview do app (que importa NOMES_ABAS), para os dois nunca divergirem.
NOMES_ABAS = (
    "Capa",
    "Premissas",
    "Modelo",
    "FCFF",
    "FCFE",
    "Macro",
    "Sensibilidades",
    "Avisos",
)

# --- Convencao de cores de Lucas (Principio 11) ---
COR_HISTORICO = "FF0B5394"  # AZUL: dado historico da CVM
COR_PREMISSA = "FF1E7A34"  # VERDE: premissa do usuario
COR_FORMULA = "FF000000"  # PRETO: formula/resultado
COR_CINZA_ND = "FF8FA3BC"

# Paleta institucional para titulos/cabecalhos (identidade visual).
COR_NAVY = "FF0A1628"
COR_AZUL_ANCORA = "FF1B4F8C"
COR_BRANCO = "FFFFFFFF"
COR_VERDE_FUNDO = "FF16A34A"
COR_VERMELHO_FUNDO = "FFDC2626"
COR_AMARELO_FUNDO = "FFB45309"

FONTE_TEXTO = "Calibri"
FONTE_NUMERO = "Consolas"

FORMATO_MILHAR = "#,##0"
FORMATO_MILHAR_2 = "#,##0.00"
FORMATO_PERCENTUAL = "0.0%"
FORMATO_PERCENTUAL_2 = "0.00%"
FORMATO_PRECO = '"R$" #,##0.00'
FORMATO_MULTIPLO = '0.00"x"'
FORMATO_DIAS = "#,##0"

LIMITE_COMPRA = 0.20
LIMITE_VENDA = -0.05

AUTOR_COMENTARIO = "DCF Automatizado"


def caminho_excel(ticker: str, raiz: Path) -> Path:
    """Caminho padrao do arquivo Excel gerado para o ticker."""
    pasta = raiz / "outputs" / "excel"
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta / f"{ticker}_dcf.xlsx"


def _valores_proximos(valor_a: Any, valor_b: Any) -> bool:
    """Compara o valor recalculado em Python com o valor do motor."""
    if not isinstance(valor_a, (int, float)) or not isinstance(valor_b, (int, float)):
        return False
    return math.isclose(float(valor_a), float(valor_b), rel_tol=1e-6, abs_tol=1e-4)


def _numero(valor: Any) -> float | None:
    """Float sem aceitar booleanos; invalido vira None."""
    if isinstance(valor, bool) or not isinstance(valor, (int, float)):
        return None
    if isinstance(valor, float) and not math.isfinite(valor):
        return None
    return float(valor)


# ---------------------------------------------------------------------------
# Contexto (tudo que o exportador consome vem persistido pelo motor)
# ---------------------------------------------------------------------------


def montar_contexto(ticker: str, raiz_projeto: Path | None = None) -> dict[str, Any]:
    """Carrega motor + premissas + historico CVM + macro para o exportador."""
    raiz = resolver_raiz(raiz_projeto)
    ticker_normalizado = normalizar_ticker(ticker)
    metadados = carregar_metadados(ticker_normalizado, raiz)
    if str(metadados.get("tipo", "")) == "financeira":
        raise RuntimeError(
            f"{ticker_normalizado} e financeira (banco/seguradora): o Excel "
            "'Modelo' cobre apenas nao-financeiras. O valuation FCFE/Ke fica "
            "no app e em data/processed/; o Excel bancario e backlog (v2.2)."
        )
    projecao = carregar_json(
        raiz / "data" / "processed" / f"{ticker_normalizado}_projecao.json"
    )
    for bloco in ("dre", "wk", "ppe", "divida", "balanco", "dfc", "fcff", "fcfe"):
        if not isinstance(projecao.get(bloco), dict):
            raise RuntimeError(
                f"Bloco obrigatorio ausente na projecao: {bloco}. Rode o "
                "pipeline completo (main.py) antes de exportar."
            )
    premissas = carregar_json(
        raiz / "data" / "premissas" / f"{ticker_normalizado}_premissas.json"
    )
    series = montar_series_anuais(ticker_normalizado, raiz)
    anos_historicos = sorted(series.get("receita_liquida", {}))

    caminho_macro = raiz / "data" / "raw" / "macro" / "macro_brasil.json"
    macro = carregar_json(caminho_macro) if caminho_macro.exists() else {}
    caminho_metricas = (
        raiz / "data" / "processed" / f"{ticker_normalizado}_metricas.json"
    )
    metricas = carregar_json(caminho_metricas) if caminho_metricas.exists() else {}
    caminho_auditoria = (
        raiz / "data" / "raw" / "cvm" / f"{ticker_normalizado}_auditoria_cvm.json"
    )
    auditoria = carregar_json(caminho_auditoria) if caminho_auditoria.exists() else {}

    return {
        "ticker": ticker_normalizado,
        "raiz": raiz,
        "projecao": projecao,
        "premissas": premissas,
        "metadados": metadados,
        "metricas": metricas,
        "series": series,
        "anos_historicos": anos_historicos,
        "macro": macro,
        "auditoria": auditoria,
        "usa_ret": empresa_usa_ret(premissas, metadados),
        # Convencao de desconto do motor (meio-periodo/stub) — as formulas do
        # Excel usam o MESMO expoente que o motor, senao o VP explicito nao
        # fecharia com os totais persistidos quando a convencao muda.
        "convencao_desconto": carregar_convencao_desconto(raiz),
    }


# ---------------------------------------------------------------------------
# Aba builder: escreve celulas rastreando valores (para o preview do app)
# ---------------------------------------------------------------------------


class Aba:
    """Worksheet + matriz de valores paralela (preview) + registro de linhas."""

    def __init__(self, wb: Workbook, nome: str) -> None:
        self.ws: Worksheet = wb.create_sheet(nome)
        self.ws.sheet_view.showGridLines = False
        self.valores: dict[tuple[int, int], Any] = {}
        self.linhas: dict[str, int] = {}
        self.cursor = 1

    def registrar(self, chave: str, linha: int) -> None:
        """Registra o numero da linha de uma chave logica."""
        self.linhas[chave] = linha

    def L(self, chave: str) -> int:  # noqa: N802 - notacao curta de layout
        """Numero da linha registrado para a chave."""
        return self.linhas[chave]

    def titulo(self, linha: int, texto: str, ate_coluna: int) -> None:
        """Faixa de titulo navy."""
        for coluna in range(1, ate_coluna + 1):
            self.ws.cell(row=linha, column=coluna).fill = PatternFill(
                "solid", start_color=COR_NAVY
            )
        celula = self.ws.cell(row=linha, column=1, value=texto)
        celula.font = Font(name=FONTE_TEXTO, bold=True, size=12, color=COR_BRANCO)
        self.valores[(linha, 1)] = texto

    def cabecalho(self, linha: int, coluna_inicial: int, rotulos: list[str]) -> None:
        """Cabecalho de colunas com fundo azul ancora."""
        for indice, rotulo in enumerate(rotulos):
            celula = self.ws.cell(
                row=linha, column=coluna_inicial + indice, value=rotulo
            )
            celula.fill = PatternFill("solid", start_color=COR_AZUL_ANCORA)
            celula.font = Font(name=FONTE_TEXTO, bold=True, size=10, color=COR_BRANCO)
            celula.alignment = Alignment(horizontal="center")
            self.valores[(linha, coluna_inicial + indice)] = rotulo

    def rotulo(
        self,
        linha: int,
        texto: str,
        negrito: bool = False,
        recuo: int = 0,
        comentario: str | None = None,
    ) -> None:
        """Rotulo da linha na coluna A (com comentario opcional)."""
        celula = self.ws.cell(row=linha, column=1, value=texto)
        celula.font = Font(name=FONTE_TEXTO, bold=negrito, size=10)
        if recuo:
            celula.alignment = Alignment(indent=recuo)
        if comentario:
            celula.comment = Comment(comentario, AUTOR_COMENTARIO, height=90, width=320)
        self.valores[(linha, 1)] = texto

    def numero(
        self,
        linha: int,
        coluna: int,
        valor: Any,
        formato: str = FORMATO_MILHAR,
        cor: str = COR_FORMULA,
        negrito: bool = False,
        comentario: str | None = None,
    ) -> None:
        """Valor numerico com fonte monoespacada; None vira 'n/d' cinza."""
        celula = self.ws.cell(row=linha, column=coluna)
        numero = _numero(valor)
        if numero is not None:
            celula.value = numero
            celula.number_format = formato
            celula.font = Font(name=FONTE_NUMERO, size=10, color=cor, bold=negrito)
            self.valores[(linha, coluna)] = numero
        else:
            celula.value = "n/d"
            celula.font = Font(name=FONTE_NUMERO, size=10, color=COR_CINZA_ND)
            celula.alignment = Alignment(horizontal="right")
            self.valores[(linha, coluna)] = None
        if comentario:
            celula.comment = Comment(comentario, AUTOR_COMENTARIO, height=90, width=320)

    def calculo(
        self,
        linha: int,
        coluna: int,
        formula: str,
        valor_python: Any,
        valor_motor: Any,
        formato: str = FORMATO_MILHAR,
        cor: str = COR_FORMULA,
        negrito: bool = False,
    ) -> None:
        """Formula nativa SE ela reproduz o valor do motor; senao o valor.

        ``valor_python`` e o resultado da formula recalculado em Python com
        os MESMOS operandos que a formula referencia. Divergiu (salvaguarda,
        piso, truncamento) => a celula recebe o valor do motor + comentario.
        """
        if _valores_proximos(valor_python, valor_motor):
            celula = self.ws.cell(row=linha, column=coluna, value=formula)
            celula.number_format = formato
            celula.font = Font(name=FONTE_NUMERO, size=10, color=cor, bold=negrito)
            self.valores[(linha, coluna)] = _numero(valor_motor)
        else:
            self.numero(linha, coluna, valor_motor, formato, cor=cor, negrito=negrito)
            if _numero(valor_motor) is not None and _numero(valor_python) is not None:
                self.ws.cell(row=linha, column=coluna).comment = Comment(
                    "Valor do motor (salvaguarda/piso impede a formula "
                    f"padrao; formula daria {float(valor_python):,.2f}).",
                    AUTOR_COMENTARIO,
                )

    def texto(
        self,
        linha: int,
        coluna: int,
        valor: str,
        cor: str = COR_FORMULA,
        negrito: bool = False,
        tamanho: int = 10,
    ) -> None:
        """Texto simples."""
        celula = self.ws.cell(row=linha, column=coluna, value=valor)
        celula.font = Font(name=FONTE_TEXTO, size=tamanho, color=cor, bold=negrito)
        self.valores[(linha, coluna)] = valor


def _ref(aba: str, coluna: int, linha: int, travar: bool = False) -> str:
    """Referencia A1 (opcionalmente absoluta) para formulas."""
    letra = get_column_letter(coluna)
    if travar:
        return f"{aba}!${letra}${linha}" if aba else f"${letra}${linha}"
    return f"{aba}!{letra}{linha}" if aba else f"{letra}{linha}"


# ---------------------------------------------------------------------------
# Helpers de dados (series, premissas, projecao)
# ---------------------------------------------------------------------------


def _proj(ctx: dict[str, Any], bloco: str, campo: str, ano: int) -> float | None:
    """Valor projetado de um campo em um bloco/ano; None quando ausente."""
    return _numero(ctx["projecao"].get(bloco, {}).get(f"ano{ano}", {}).get(campo))


def _hist(ctx: dict[str, Any], serie: str, ano: int) -> float | None:
    """Valor historico CVM de uma serie padronizada em um exercicio."""
    return _numero(ctx["series"].get(serie, {}).get(ano))


def _vetor_premissa(premissas: dict[str, Any], base: str) -> list[float | None]:
    """Vetor ano1..8 de uma premissa (None quando ausente)."""
    return [
        _numero(premissas.get(f"{base}_ano{ano}"))
        for ano in range(1, HORIZONTE_PROJECAO + 1)
    ]


# ---------------------------------------------------------------------------
# Aba 1 — Capa
# ---------------------------------------------------------------------------


def _aba_capa(wb: Workbook, ctx: dict[str, Any]) -> Aba:
    """Capa: identificacao, decisao e a LEGENDA DE CORES de Lucas."""
    aba = Aba(wb, "Capa")
    ws = aba.ws
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40

    meta = ctx["metadados"]
    ev = ctx["projecao"].get("ev_equity", {})
    aba.titulo(1, "DCF AUTOMATIZADO — MODELO DE VALUATION", 8)
    linhas = [
        ("Empresa", str(meta.get("razao_social", ctx["ticker"]))),
        ("Ticker", ctx["ticker"]),
        (
            "Setor / subtipo",
            f"{meta.get('setor', 'n/d')} / {meta.get('subtipo', 'n/d')}",
        ),
        (
            "Data-base (Ano 0)",
            str(ctx["anos_historicos"][-1]) if ctx["anos_historicos"] else "n/d",
        ),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Metodo", "DCF FCFF/WACC (bridge) + FCFE/Ke (checagem)"),
    ]
    linha = 3
    for rotulo, valor in linhas:
        aba.rotulo(linha, rotulo, negrito=True)
        aba.texto(linha, 2, valor)
        linha += 1

    linha += 1
    aba.titulo(linha, "DECISAO", 8)
    linha += 1
    target = _numero(ev.get("target_price"))
    preco = _numero(ev.get("preco_atual"))
    upside = _numero(ev.get("upside"))
    aba.rotulo(linha, "Target Price (bridge FCFF)", negrito=True)
    aba.numero(linha, 2, target, FORMATO_PRECO, negrito=True)
    linha += 1
    aba.rotulo(linha, "Preco atual")
    aba.numero(linha, 2, preco, FORMATO_PRECO)
    linha += 1
    aba.rotulo(linha, "Upside")
    aba.numero(linha, 2, upside, FORMATO_PERCENTUAL)
    linha += 1
    aba.rotulo(linha, "Recomendacao", negrito=True)
    aba.texto(linha, 2, str(ev.get("recomendacao", "n/d")), negrito=True)
    linha += 2

    # Legenda de cores — a convencao de Lucas (obrigatoria na Capa).
    aba.titulo(linha, "LEGENDA DE CORES", 8)
    linha += 1
    aba.rotulo(linha, "Historico (dado da CVM)")
    aba.texto(linha, 2, "AZUL", cor=COR_HISTORICO, negrito=True)
    aba.registrar("legenda_historico", linha)
    linha += 1
    aba.rotulo(linha, "Premissa (voce escolhe)")
    aba.texto(linha, 2, "VERDE", cor=COR_PREMISSA, negrito=True)
    aba.registrar("legenda_premissa", linha)
    linha += 1
    aba.rotulo(linha, "Formula (resultado)")
    aba.texto(linha, 2, "PRETO", cor=COR_FORMULA, negrito=True)
    aba.registrar("legenda_formula", linha)
    linha += 2

    if ctx["premissas"].get("premissas_automaticas"):
        aba.texto(
            linha,
            1,
            "ATENCAO: PREMISSAS AUTOMATICAS DE PARTIDA (ancoras historicas + "
            "defaults do subtipo). A tese e do analista — revise as celulas "
            "VERDES antes de usar como recomendacao.",
            cor=COR_VERMELHO_FUNDO,
            negrito=True,
        )
    return aba


# ---------------------------------------------------------------------------
# Aba 2 — Premissas (as 6 de Lucas, em VERDE, com ancora e comentario)
# ---------------------------------------------------------------------------

VETORES_PREMISSAS = (
    (
        "crescimento_receita",
        "1. Crescimento da receita liquida",
        FORMATO_PERCENTUAL_2,
        "Crescimento % anual da receita liquida (excecao ao Direcional: sem "
        "build-up VGVxPOC).",
        "cagr_receita_3a",
    ),
    (
        "margem_bruta",
        "2. Margem bruta (nivel EBITDA, pre-D&A)",
        FORMATO_PERCENTUAL_2,
        "Margem bruta PRE-D&A: o CPV nao carrega depreciacao; a D&A e linha "
        "propria subtraida depois (padrao Direcional).",
        "margem_bruta_media_3a",
    ),
    (
        "sgna_pct_receita",
        "3. SG&A % da receita liquida",
        FORMATO_PERCENTUAL_2,
        "Despesas comerciais + G&A como % da receita liquida (pre-D&A).",
        None,
    ),
    (
        "aliquota_ir",
        "4. Aliquota IR/CSLL anual",
        FORMATO_PERCENTUAL_2,
        "Aliquota sobre o EBT positivo. Construtora RET ignora este vetor "
        "(4% sobre a Receita Bruta).",
        "aliquota_efetiva_media_3a",
    ),
    (
        "deducoes_pct_receita_bruta",
        "Deducoes % da receita bruta",
        FORMATO_PERCENTUAL_2,
        "Deducoes (impostos s/ vendas) como % da receita bruta.",
        None,
    ),
    (
        "capex_receita",
        "CAPEX % da receita (negativo = saida)",
        FORMATO_PERCENTUAL_2,
        "CAPEX como % da receita liquida; negativo por convencao de sinal.",
        "capex_receita_media_3a",
    ),
)

ESCALARES_PREMISSAS = (
    (
        "crescimento_perpetuidade_g",
        "g — crescimento na perpetuidade",
        FORMATO_PERCENTUAL_2,
        "Crescimento nominal de longo prazo (Gordon).",
    ),
    (
        "custo_divida_kd",
        "Kd — custo da divida",
        FORMATO_PERCENTUAL_2,
        "Custo medio da divida nova/rolada.",
    ),
    (
        "beta",
        "Beta desalavancado",
        FORMATO_MILHAR_2,
        "Beta desalavancado; re-alavancado por Hamada no WACC.",
    ),
    (
        "wacc_manual",
        "5. WACC manual (opcional)",
        FORMATO_PERCENTUAL_2,
        "Se preenchido, o motor usa ESTE WACC no lugar do build-up CAPM.",
    ),
    (
        "payout_dividendos",
        "Payout de dividendos",
        FORMATO_PERCENTUAL,
        "Percentual do lucro liquido distribuido.",
    ),
    (
        "minoritarios_pct_ll",
        "Minoritarios % do LL",
        FORMATO_PERCENTUAL_2,
        "Participacao de nao controladores sobre o LL.",
    ),
    (
        "outras_despesas_pct_receita",
        "Outras receitas/despesas % RL",
        FORMATO_PERCENTUAL_2,
        "Outras operacionais como % da receita liquida.",
    ),
    (
        "equivalencia_pct_receita",
        "Equivalencia patrimonial % RL",
        FORMATO_PERCENTUAL_2,
        "Resultado de equivalencia como % da receita.",
    ),
    (
        "caixa_minimo_pct_receita",
        "Caixa minimo % da receita",
        FORMATO_PERCENTUAL_2,
        "Piso de caixa; deficit vira captacao automatica.",
    ),
)


def _aba_premissas(wb: Workbook, ctx: dict[str, Any]) -> Aba:
    """Premissas em VERDE (vetores x8 + escalares) com ancora historica."""
    aba = Aba(wb, "Premissas")
    ws = aba.ws
    ws.column_dimensions["A"].width = 40
    for coluna in range(2, 10):
        ws.column_dimensions[get_column_letter(coluna)].width = 11
    ws.column_dimensions["K"].width = 42

    premissas = ctx["premissas"]
    agregados = ctx["metricas"].get("agregados", {})
    aba.titulo(1, "PREMISSAS DO ANALISTA (celulas VERDES = voce escolhe)", 11)
    aba.cabecalho(2, 2, [f"Ano {ano}" for ano in range(1, HORIZONTE_PROJECAO + 1)])
    aba.texto(2, 11, "Ancora historica", cor=COR_CINZA_ND)

    linha = 3
    for base, rotulo, formato, comentario, chave_ancora in VETORES_PREMISSAS:
        if base == "aliquota_ir" and ctx["usa_ret"]:
            aba.rotulo(linha, rotulo, comentario=comentario)
            aba.texto(
                linha,
                2,
                "RET: 4% sobre a Receita Bruta projetada (travado pelo motor)",
                cor=COR_CINZA_ND,
            )
            aba.registrar("aliquota_ir", linha)
            linha += 1
            continue
        aba.rotulo(linha, rotulo, comentario=comentario)
        vetor = _vetor_premissa(premissas, base)
        for indice, valor in enumerate(vetor):
            # Premissa ausente vira 0,0 VERDE (nunca "n/d"): as formulas do
            # Modelo referenciam estas celulas e precisam de numero.
            aba.numero(
                linha,
                2 + indice,
                valor if valor is not None else 0.0,
                formato,
                cor=COR_PREMISSA,
            )
        ancora = _numero(agregados.get(chave_ancora)) if chave_ancora else None
        if ancora is not None:
            aba.texto(linha, 11, f"hist. 3a: {ancora * 100:,.1f}%", cor=COR_CINZA_ND)
        aba.registrar(base, linha)
        linha += 1

    linha += 1
    aba.titulo(linha, "ESCALARES", 11)
    linha += 1
    for campo, rotulo, formato, comentario in ESCALARES_PREMISSAS:
        valor = _numero(premissas.get(campo))
        if campo == "wacc_manual" and valor is None:
            aba.rotulo(linha, rotulo, comentario=comentario)
            aba.texto(linha, 2, "vazio = build-up CAPM (aba FCFF)", cor=COR_CINZA_ND)
            aba.registrar(campo, linha)
            linha += 1
            continue
        aba.rotulo(linha, rotulo, comentario=comentario)
        # Escalar ausente vira 0,0 (formulas do Modelo referenciam a celula).
        aba.numero(
            linha, 2, valor if valor is not None else 0.0, formato, cor=COR_PREMISSA
        )
        aba.registrar(campo, linha)
        linha += 1

    aba.texto(
        linha + 1,
        1,
        "Comentarios de celula (passe o mouse) explicam cada premissa. "
        "Vetores tem 8 valores INDIVIDUAIS por regra do projeto.",
        cor=COR_CINZA_ND,
    )
    return aba


# ---------------------------------------------------------------------------
# Aba 3 — Modelo (DRE + BP aberto + DFC indireto + WK + Divida + PP&E)
# ---------------------------------------------------------------------------


def _layout_modelo(ctx: dict[str, Any]) -> list[tuple[str, str]]:
    """Ordem das linhas do Modelo: lista de (chave, tipo).

    tipo: 'titulo' | 'cab' (cabecalho de anos) | 'linha' | 'vazio'.
    Blocos condicionais (leasing, contas do WK) entram conforme o motor.
    """
    projecao = ctx["projecao"]
    tem_leasing = bool(
        _numero(projecao.get("dre", {}).get("ano1", {}).get("juros_arrendamento"))
    )
    wk_ano1 = projecao.get("wk", {}).get("ano1", {})
    contas_wk = [
        ("wk_contas_receber", "contas_receber"),
        ("wk_estoques", "estoques"),
        ("wk_tributos", "tributos_a_recuperar"),
        ("wk_fornecedores", "fornecedores"),
        ("wk_obrigacoes", "obrigacoes_sociais_trabalhistas"),
        ("wk_adiantamentos", "adiantamento_clientes"),
    ]
    contas_wk_presentes = [chave for chave, campo in contas_wk if campo in wk_ano1]

    layout: list[tuple[str, str]] = [
        ("titulo_modelo", "titulo"),
        ("cab_anos", "cab"),
        ("sec_premissas", "titulo"),
        ("p_crescimento", "linha"),
        ("p_margem_bruta", "linha"),
        ("p_sgna", "linha"),
        ("p_deducoes", "linha"),
        ("p_aliquota", "linha"),
        ("p_capex", "linha"),
        ("vazio1", "vazio"),
        ("sec_dre", "titulo"),
        ("receita_bruta", "linha"),
        ("deducoes", "linha"),
        ("receita_liquida", "linha"),
        ("cpv", "linha"),
        ("lucro_bruto", "linha"),
        ("memo_margem_bruta", "linha"),
        ("sgna", "linha"),
        ("outras", "linha"),
        ("equivalencia", "linha"),
        ("ebit_ex_da", "linha"),
        ("dre_da", "linha"),
        ("ebit", "linha"),
        ("memo_margem_ebit", "linha"),
        ("resultado_financeiro", "linha"),
        ("ebt", "linha"),
        ("ir_csll", "linha"),
        ("ll_antes_min", "linha"),
        ("minoritarios", "linha"),
        ("lucro_liquido", "linha"),
        ("memo_margem_liquida", "linha"),
        ("lpa", "linha"),
        ("vazio2", "vazio"),
        ("sec_bp", "titulo"),
        ("bp_caixa", "linha"),
        ("bp_aplicacoes", "linha"),
        ("bp_contas_receber", "linha"),
        ("bp_estoques", "linha"),
        ("bp_tributos", "linha"),
        ("bp_imobilizado", "linha"),
        ("bp_intangivel", "linha"),
        ("bp_outros_ativos", "linha"),
        ("bp_ativo_total", "linha"),
        ("vazio3", "vazio"),
        ("bp_fornecedores", "linha"),
        ("bp_obrigacoes", "linha"),
        ("bp_adiantamentos", "linha"),
        ("bp_divida_cp", "linha"),
        ("bp_divida_lp", "linha"),
        ("bp_arrendamento", "linha"),
        ("bp_outros_passivos", "linha"),
        ("bp_passivo_total", "linha"),
        ("bp_pl", "linha"),
        ("bp_passivo_pl", "linha"),
        ("bp_check", "linha"),
        ("vazio4", "vazio"),
        ("sec_dfc", "titulo"),
        ("dfc_ll", "linha"),
        ("dfc_da", "linha"),
    ]
    for chave in contas_wk_presentes:
        layout.append((f"dfc_var_{chave}", "linha"))
    layout += [
        ("dfc_fco", "linha"),
        ("dfc_capex", "linha"),
        ("dfc_fci", "linha"),
        ("dfc_dividendos", "linha"),
        ("dfc_delta_divida", "linha"),
        ("dfc_fcfin", "linha"),
        ("dfc_caixa_bop", "linha"),
        ("dfc_variacao", "linha"),
        ("dfc_caixa_eop", "linha"),
        ("vazio5", "vazio"),
        ("sec_wk", "titulo"),
    ]
    for chave in contas_wk_presentes:
        layout.append((chave, "linha"))
        layout.append((f"{chave}_dias", "linha"))
    layout += [
        ("wk_nwc", "linha"),
        ("wk_delta_nwc", "linha"),
        ("vazio6", "vazio"),
        ("sec_divida", "titulo"),
        ("div_bop", "linha"),
        ("div_captacao", "linha"),
        ("div_amortizacao", "linha"),
        ("div_eop", "linha"),
        ("div_cp", "linha"),
        ("div_lp", "linha"),
        ("div_kd", "linha"),
        ("div_juros", "linha"),
        ("div_taxa_aplicacao", "linha"),
        ("div_receita_financeira", "linha"),
    ]
    if tem_leasing:
        layout += [
            ("vazio7", "vazio"),
            ("sec_leasing", "titulo"),
            ("leasing_passivo", "linha"),
            ("leasing_juros", "linha"),
            ("leasing_da_direito_uso", "linha"),
        ]
    layout += [
        ("vazio8", "vazio"),
        ("sec_ppe", "titulo"),
        ("ppe_bop", "linha"),
        ("ppe_capex_pct", "linha"),
        ("ppe_capex", "linha"),
        ("ppe_da_pct", "linha"),
        ("ppe_da", "linha"),
        ("ppe_eop", "linha"),
        ("ppe_da_pct_historica", "linha"),
        ("ppe_intangivel", "linha"),
    ]
    ctx["_contas_wk_presentes"] = contas_wk_presentes
    ctx["_tem_leasing"] = tem_leasing
    return layout


CONTAS_WK_INFO = {
    "wk_contas_receber": (
        "Contas a receber",
        "contas_receber",
        "contas_receber",
        "receita_liquida",
        "Dias de Receita Liquida",
    ),
    "wk_estoques": ("Estoques", "estoques", "estoques", "cpv", "Dias de CPV"),
    "wk_tributos": (
        "Tributos a recuperar",
        "tributos_a_recuperar",
        "tributos_a_recuperar",
        "receita_liquida",
        "Dias (driver do motor)",
    ),
    "wk_fornecedores": (
        "Fornecedores",
        "fornecedores",
        "fornecedores",
        "cpv",
        "Dias de CPV",
    ),
    "wk_obrigacoes": (
        "Obrigacoes sociais e trabalhistas",
        "obrigacoes_sociais_trabalhistas",
        "obrigacoes_sociais_trabalhistas",
        "sgna",
        "Dias de SG&A",
    ),
    "wk_adiantamentos": (
        "Adiantamento de clientes",
        "adiantamento_clientes",
        None,
        "receita_liquida",
        "Dias de Receita Liquida",
    ),
}


def _aba_modelo(wb: Workbook, ctx: dict[str, Any]) -> Aba:
    """A espinha dorsal: 3 demonstrativos + schedules, historico + projecao."""
    aba = Aba(wb, "Modelo")
    ws = aba.ws
    layout = _layout_modelo(ctx)
    # Pre-calcula o numero de cada linha (permite referencias para frente,
    # como a Direcional: DRE L28 -> PP&E L199).
    for indice, (chave, _) in enumerate(layout, start=1):
        aba.registrar(chave, indice)

    anos_hist = ctx["anos_historicos"]
    nh = len(anos_hist)
    col_h0 = 2  # primeira coluna historica
    col_p0 = 2 + nh  # primeira coluna projetada (Ano 1)

    def CP(t: int) -> int:  # noqa: N802 - coluna projetada do ano t
        return col_p0 + (t - 1)

    def CH(i: int) -> int:  # noqa: N802 - coluna historica de indice i
        return col_h0 + i

    def rp(chave: str, t: int, travar: bool = False) -> str:
        """Referencia local (mesma aba) a uma linha em uma coluna projetada."""
        return _ref("", CP(t), aba.L(chave), travar)

    ws.freeze_panes = _ref("", col_p0, 3).replace("$", "")
    ws.column_dimensions["A"].width = 40
    for coluna in range(2, col_p0 + HORIZONTE_PROJECAO):
        ws.column_dimensions[get_column_letter(coluna)].width = 13

    premissas = ctx["premissas"]
    usa_ret = ctx["usa_ret"]

    def p(bloco: str, campo: str, t: int) -> float | None:
        return _proj(ctx, bloco, campo, t)

    def h(nome: str, i: int) -> float | None:
        return _hist(ctx, nome, anos_hist[i])

    # Referencias a aba Premissas (linhas registradas na propria aba).
    aba_premissas: Aba = ctx["_aba_premissas"]

    def prem(base: str, t: int, travar: bool = False) -> str:
        coluna = 1 + t if t else 2
        return _ref("Premissas", coluna, aba_premissas.L(base), travar)

    # --- Escrita ---
    aba.titulo(aba.L("titulo_modelo"), f"MODELO — {ctx['ticker']} (R$ mil)", col_p0 + 8)
    rotulos_cab = [str(ano) for ano in anos_hist] + [
        f"Ano {t}" for t in range(1, HORIZONTE_PROJECAO + 1)
    ]
    aba.cabecalho(aba.L("cab_anos"), 2, rotulos_cab)

    # ---- Premissas do topo (referencias VERDES a aba Premissas) ----
    aba.titulo(
        aba.L("sec_premissas"),
        "PREMISSAS (referencias — edite na aba Premissas)",
        col_p0 + 8,
    )
    topo = (
        (
            "p_crescimento",
            "Crescimento da receita (%)",
            "crescimento_receita",
            "taxa_crescimento_receita",
        ),
        ("p_margem_bruta", "Margem bruta (%)", "margem_bruta", "margem_bruta"),
        ("p_sgna", "SG&A % receita", "sgna_pct_receita", "sgna_pct_receita"),
        ("p_deducoes", "Deducoes % receita bruta", "deducoes_pct_receita_bruta", None),
        ("p_aliquota", "Aliquota IR/CSLL", "aliquota_ir", "aliquota_ir_ano"),
        ("p_capex", "CAPEX % receita", "capex_receita", None),
    )
    for chave, rotulo, base, campo_dre in topo:
        linha = aba.L(chave)
        aba.rotulo(linha, rotulo, recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            if chave == "p_aliquota" and usa_ret:
                if t == 1:
                    aba.texto(linha, CP(t), "RET 4% s/ RB", cor=COR_CINZA_ND)
                continue
            if chave == "p_deducoes":
                valor_motor = _numero(premissas.get(f"{base}_ano{t}")) or 0.0
            elif chave == "p_capex":
                valor_motor = _numero(premissas.get(f"{base}_ano{t}"))
            elif campo_dre:
                valor_motor = p("dre", campo_dre, t)
                if valor_motor is None:
                    valor_motor = _numero(premissas.get(f"{base}_ano{t}"))
            else:
                valor_motor = _numero(premissas.get(f"{base}_ano{t}"))
            valor_premissa = _numero(premissas.get(f"{base}_ano{t}"))
            if chave == "p_deducoes" and valor_premissa is None:
                valor_premissa = 0.0
            # Referencia direta a premissa => celula VERDE (convencao).
            aba.calculo(
                linha,
                CP(t),
                f"={prem(base, t)}",
                valor_premissa,
                valor_motor,
                FORMATO_PERCENTUAL_2,
                cor=COR_PREMISSA,
            )

    # ---- DRE ----
    aba.titulo(
        aba.L("sec_dre"), "DRE PROJETADA (pre-D&A, padrao Direcional)", col_p0 + 8
    )

    def linha_dre(
        chave: str,
        rotulo: str,
        campo: str,
        serie_hist: str | None,
        formula_fn: Callable[[int], tuple[str, float | None]] | None,
        formato: str = FORMATO_MILHAR,
        negrito: bool = False,
        recuo: int = 1,
        hist_fn: Callable[[int], float | None] | None = None,
        sinal_motor: float = 1.0,
    ) -> None:
        """Linha da DRE; ``sinal_motor`` ajusta a EXIBICAO (ex.: D&A negativa)."""
        linha = aba.L(chave)
        aba.rotulo(linha, rotulo, negrito=negrito, recuo=recuo)
        for i in range(nh):
            valor_hist = (
                hist_fn(i) if hist_fn else (h(serie_hist, i) if serie_hist else None)
            )
            aba.numero(linha, CH(i), valor_hist, formato, cor=COR_HISTORICO)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            bruto = p("dre", campo, t) if campo else None
            valor_motor = bruto * sinal_motor if bruto is not None else None
            if campo is None and formula_fn is not None:
                # Linha memo (margens %): formula de apresentacao sempre.
                formula, valor_python = formula_fn(t)
                aba.calculo(
                    linha,
                    CP(t),
                    formula,
                    valor_python,
                    valor_python,
                    formato,
                    negrito=negrito,
                )
                continue
            if formula_fn is None or valor_motor is None:
                aba.numero(linha, CP(t), valor_motor, formato)
                continue
            formula, valor_python = formula_fn(t)
            aba.calculo(
                linha,
                CP(t),
                formula,
                valor_python,
                valor_motor,
                formato,
                negrito=negrito,
            )

    def v(campo: str, t: int) -> float | None:
        return p("dre", campo, t)

    def f_receita_bruta(t: int) -> tuple[str, float | None]:
        rl = v("receita_liquida", t)
        ded = _numero(premissas.get(f"deducoes_pct_receita_bruta_ano{t}")) or 0.0
        formula = f"={rp('receita_liquida', t)}/(1-{rp('p_deducoes', t)})"
        return formula, (rl / (1 - ded) if rl is not None and ded < 1 else None)

    def f_deducoes(t: int) -> tuple[str, float | None]:
        rl, rb = v("receita_liquida", t), v("receita_bruta", t)
        return (
            f"={rp('receita_liquida', t)}-{rp('receita_bruta', t)}",
            (rl - rb) if rl is not None and rb is not None else None,
        )

    def f_receita_liquida(t: int) -> tuple[str, float | None]:
        cresc = v("taxa_crescimento_receita", t)
        if t == 1:
            base = h("receita_liquida", nh - 1) if nh else None
            ref_base = _ref("", CH(nh - 1), aba.L("receita_liquida"))
        else:
            base = v("receita_liquida", t - 1)
            ref_base = rp("receita_liquida", t - 1)
        formula = f"={ref_base}*(1+{rp('p_crescimento', t)})"
        valor = base * (1 + cresc) if base is not None and cresc is not None else None
        return formula, valor

    def f_cpv(t: int) -> tuple[str, float | None]:
        rl, mb = v("receita_liquida", t), v("margem_bruta", t)
        formula = f"=-{rp('receita_liquida', t)}*(1-{rp('p_margem_bruta', t)})"
        return formula, (-rl * (1 - mb) if rl is not None and mb is not None else None)

    def f_lucro_bruto(t: int) -> tuple[str, float | None]:
        rl, cpv = v("receita_liquida", t), v("cpv_cmv", t)
        return (
            f"={rp('receita_liquida', t)}+{rp('cpv', t)}",
            (rl + cpv) if rl is not None and cpv is not None else None,
        )

    def f_sgna(t: int) -> tuple[str, float | None]:
        rl, pct = v("receita_liquida", t), v("sgna_pct_receita", t)
        formula = f"=-{rp('receita_liquida', t)}*{rp('p_sgna', t)}"
        return formula, (-rl * pct if rl is not None and pct is not None else None)

    def f_outras(t: int) -> tuple[str, float | None]:
        rl = v("receita_liquida", t)
        pct = _numero(premissas.get("outras_despesas_pct_receita")) or 0.0
        formula = (
            f"={rp('receita_liquida', t)}*"
            f"{prem('outras_despesas_pct_receita', 0, travar=True)}"
        )
        return formula, (rl * pct if rl is not None else None)

    def f_equivalencia(t: int) -> tuple[str, float | None]:
        rl = v("receita_liquida", t)
        pct = _numero(premissas.get("equivalencia_pct_receita")) or 0.0
        formula = (
            f"={rp('receita_liquida', t)}*"
            f"{prem('equivalencia_pct_receita', 0, travar=True)}"
        )
        return formula, (rl * pct if rl is not None else None)

    def f_ebit_ex_da(t: int) -> tuple[str, float | None]:
        partes = [
            v("lucro_bruto", t),
            v("sgna", t),
            v("outras_receitas_despesas", t),
            v("equivalencia_patrimonial", t),
        ]
        formula = (
            f"={rp('lucro_bruto', t)}+{rp('sgna', t)}+{rp('outras', t)}"
            f"+{rp('equivalencia', t)}"
        )
        soma = (
            sum(x for x in partes if x is not None) if partes[0] is not None else None
        )
        return formula, soma

    def f_dre_da(t: int) -> tuple[str, float | None]:
        da = v("depreciacao_amortizacao", t)
        return f"={rp('ppe_da', t)}", (-da if da is not None else None)

    def f_ebit(t: int) -> tuple[str, float | None]:
        ex_da, da = v("ebit_ex_depreciacao", t), v("depreciacao_amortizacao", t)
        formula = f"={rp('ebit_ex_da', t)}+{rp('dre_da', t)}"
        return formula, (ex_da - da if ex_da is not None and da is not None else None)

    def f_resultado_financeiro(t: int) -> tuple[str, float | None]:
        recfin = p("divida", "receita_financeira_caixa", t)
        juros = p("divida", "juros", t)
        juros_arr = p("divida", "juros_arrendamento", t) or 0.0
        partes = f"={rp('div_receita_financeira', t)}+{rp('div_juros', t)}"
        if ctx["_tem_leasing"]:
            partes += f"+{rp('leasing_juros', t)}"
        valor = (
            recfin - juros - juros_arr
            if recfin is not None and juros is not None
            else None
        )
        return partes, valor

    def f_ebt(t: int) -> tuple[str, float | None]:
        ebit, rf = v("ebit", t), v("resultado_financeiro", t)
        return (
            f"={rp('ebit', t)}+{rp('resultado_financeiro', t)}",
            (ebit + rf) if ebit is not None and rf is not None else None,
        )

    def f_ir(t: int) -> tuple[str, float | None]:
        ebt = v("ebt", t)
        if usa_ret:
            rb = v("receita_bruta", t)
            formula = f"=-{ALIQUOTA_RET_RECEITA}*{rp('receita_bruta', t)}"
            return formula, (-ALIQUOTA_RET_RECEITA * rb if rb is not None else None)
        aliq = v("aliquota_ir_ano", t)
        formula = f"=IF({rp('ebt', t)}>0,-{rp('p_aliquota', t)}*{rp('ebt', t)},0)"
        if ebt is None or aliq is None:
            return formula, None
        return formula, (-aliq * ebt if ebt > 0 else 0.0)

    def f_ll_antes(t: int) -> tuple[str, float | None]:
        ebt, ir = v("ebt", t), v("ir_csll", t)
        return (
            f"={rp('ebt', t)}+{rp('ir_csll', t)}",
            (ebt + ir) if ebt is not None and ir is not None else None,
        )

    def f_minoritarios(t: int) -> tuple[str, float | None]:
        ll_antes = v("ll_antes_minoritarios", t)
        pct = _numero(premissas.get("minoritarios_pct_ll")) or 0.0
        formula = (
            f"=-{prem('minoritarios_pct_ll', 0, travar=True)}*{rp('ll_antes_min', t)}"
        )
        return formula, (-pct * ll_antes if ll_antes is not None else None)

    def f_ll(t: int) -> tuple[str, float | None]:
        ll_antes, minor = v("ll_antes_minoritarios", t), v(
            "participacao_minoritarios", t
        )
        return (
            f"={rp('ll_antes_min', t)}+{rp('minoritarios', t)}",
            (ll_antes + minor) if ll_antes is not None and minor is not None else None,
        )

    def memo_pct(numerador: str, t: int) -> tuple[str, float | None]:
        num, rl = v(numerador, t), v("receita_liquida", t)
        chave_num = {
            "lucro_bruto": "lucro_bruto",
            "ebit": "ebit",
            "lucro_liquido": "lucro_liquido",
        }[numerador]
        return (
            f"={rp(chave_num, t)}/{rp('receita_liquida', t)}",
            (num / rl if num is not None and rl not in (None, 0) else None),
        )

    def hist_deducoes(i: int) -> float | None:
        rl, rb = h("receita_liquida", i), h("receita_bruta", i)
        return rl - rb if rl is not None and rb is not None else None

    def hist_ebit_ex_da(i: int) -> float | None:
        ebit, da = h("ebit", i), h("depreciacao_amortizacao", i)
        return ebit + abs(da) if ebit is not None and da is not None else None

    def hist_margem(numerador: str) -> Callable[[int], float | None]:
        def interna(i: int) -> float | None:
            num, rl = h(numerador, i), h("receita_liquida", i)
            return num / rl if num is not None and rl not in (None, 0) else None

        return interna

    linha_dre(
        "receita_bruta",
        "(=) Receita Bruta",
        "receita_bruta",
        "receita_bruta",
        f_receita_bruta,
        negrito=True,
    )
    linha_dre(
        "deducoes", "(-) Deducoes", "deducoes", None, f_deducoes, hist_fn=hist_deducoes
    )
    linha_dre(
        "receita_liquida",
        "(=) Receita Liquida",
        "receita_liquida",
        "receita_liquida",
        f_receita_liquida,
        negrito=True,
    )
    linha_dre("cpv", "(-) CPV/CMV", "cpv_cmv", "cpv_cmv", f_cpv)
    linha_dre(
        "lucro_bruto",
        "(=) Lucro Bruto",
        "lucro_bruto",
        "lucro_bruto",
        f_lucro_bruto,
        negrito=True,
    )
    linha_dre(
        "memo_margem_bruta",
        "Margem bruta (%)",
        "margem_bruta",
        None,
        lambda t: memo_pct("lucro_bruto", t),
        FORMATO_PERCENTUAL,
        recuo=2,
        hist_fn=hist_margem("lucro_bruto"),
    )
    linha_dre("sgna", "(-) SG&A (comerciais + G&A)", "sgna", None, f_sgna)
    linha_dre(
        "outras",
        "(+/-) Outras receitas/despesas",
        "outras_receitas_despesas",
        None,
        f_outras,
    )
    linha_dre(
        "equivalencia",
        "(+) Equivalencia patrimonial",
        "equivalencia_patrimonial",
        None,
        f_equivalencia,
    )
    linha_dre(
        "ebit_ex_da",
        "(=) EBIT ex-Depreciacao (EBITDA)",
        "ebit_ex_depreciacao",
        None,
        f_ebit_ex_da,
        negrito=True,
        hist_fn=hist_ebit_ex_da,
    )
    linha_dre(
        "dre_da",
        "(-) Depreciacao e Amortizacao",
        "depreciacao_amortizacao",
        None,
        f_dre_da,
        sinal_motor=-1.0,
        hist_fn=lambda i: (
            -abs(h("depreciacao_amortizacao", i))
            if h("depreciacao_amortizacao", i) is not None
            else None
        ),
    )
    linha_dre("ebit", "(=) EBIT", "ebit", "ebit", f_ebit, negrito=True)
    linha_dre(
        "memo_margem_ebit",
        "Margem EBIT (%)",
        None,
        None,
        lambda t: memo_pct("ebit", t),
        FORMATO_PERCENTUAL,
        recuo=2,
        hist_fn=hist_margem("ebit"),
    )
    linha_dre(
        "resultado_financeiro",
        "(+/-) Resultado financeiro",
        "resultado_financeiro",
        None,
        f_resultado_financeiro,
        hist_fn=lambda i: h("despesas_financeiras", i),
    )
    linha_dre("ebt", "(=) EBT", "ebt", "ebt", f_ebt, negrito=True)
    linha_dre("ir_csll", "(-) IR / CSLL", "ir_csll", "ir_csll", f_ir)
    linha_dre(
        "ll_antes_min",
        "(=) LL antes de minoritarios",
        "ll_antes_minoritarios",
        None,
        f_ll_antes,
    )
    linha_dre(
        "minoritarios",
        "(-) Participacao de minoritarios",
        "participacao_minoritarios",
        None,
        f_minoritarios,
    )
    linha_dre(
        "lucro_liquido",
        "(=) Lucro Liquido",
        "lucro_liquido",
        "lucro_liquido",
        f_ll,
        negrito=True,
    )
    linha_dre(
        "memo_margem_liquida",
        "Margem liquida (%)",
        None,
        None,
        lambda t: memo_pct("lucro_liquido", t),
        FORMATO_PERCENTUAL,
        recuo=2,
        hist_fn=hist_margem("lucro_liquido"),
    )

    linha = aba.L("lpa")
    aba.rotulo(linha, "LPA (R$/acao, escala DFs)", recuo=2)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        aba.numero(linha, CP(t), v("lpa", t), FORMATO_MILHAR_2)

    # Campo memo "margem_bruta" tambem existe na dre; sem problema.

    # ---- Balanco patrimonial ABERTO ----
    aba.titulo(aba.L("sec_bp"), "BALANCO PATRIMONIAL PROJETADO (aberto)", col_p0 + 8)

    def b(campo: str, t: int) -> float | None:
        return p("balanco", campo, t)

    def linha_bp(
        chave: str,
        rotulo: str,
        campo: str,
        serie_hist: str | None,
        formula_fn: Callable[[int], tuple[str, float | None]] | None,
        negrito: bool = False,
        hist_fn: Callable[[int], float | None] | None = None,
        comentario: str | None = None,
    ) -> None:
        linha = aba.L(chave)
        aba.rotulo(linha, rotulo, negrito=negrito, recuo=1, comentario=comentario)
        for i in range(nh):
            valor_hist = (
                hist_fn(i)
                if hist_fn
                else (
                    abs(h(serie_hist, i))
                    if serie_hist and h(serie_hist, i) is not None
                    else (h(serie_hist, i) if serie_hist else None)
                )
            )
            aba.numero(linha, CH(i), valor_hist, FORMATO_MILHAR, cor=COR_HISTORICO)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            valor_motor = b(campo, t)
            if formula_fn is None or valor_motor is None:
                aba.numero(linha, CP(t), valor_motor, FORMATO_MILHAR, negrito=negrito)
                continue
            formula, valor_python = formula_fn(t)
            aba.calculo(
                linha,
                CP(t),
                formula,
                valor_python,
                valor_motor,
                FORMATO_MILHAR,
                negrito=negrito,
            )

    def f_constante(chave: str, serie_hist: str | None, campo: str):
        def interna(t: int) -> tuple[str, float | None]:
            if t == 1:
                valor_base = (
                    abs(h(serie_hist, nh - 1))
                    if serie_hist and h(serie_hist, nh - 1) is not None
                    else None
                )
                ref_base = _ref("", CH(nh - 1), aba.L(chave))
            else:
                valor_base = b(campo, t - 1)
                ref_base = rp(chave, t - 1)
            return f"={ref_base}", valor_base

        return interna

    def hist_outros_ativos(i: int) -> float | None:
        total = h("ativo_total", i)
        if total is None:
            return None
        mapeados = 0.0
        for nome in (
            "caixa_equivalentes",
            "aplicacoes_financeiras",
            "contas_receber",
            "estoques",
            "tributos_a_recuperar",
            "imobilizado",
            "intangivel",
        ):
            valor = h(nome, i)
            if valor is not None:
                mapeados += abs(valor)
        return total - mapeados

    def hist_outros_passivos(i: int) -> float | None:
        total, pl = h("passivo_total", i), h("patrimonio_liquido", i)
        if total is None or pl is None:
            return None
        exigivel = total - pl
        mapeados = 0.0
        for nome in (
            "fornecedores",
            "obrigacoes_sociais_trabalhistas",
            "divida_curto_prazo",
            "divida_longo_prazo",
        ):
            valor = h(nome, i)
            if valor is not None:
                mapeados += abs(valor)
        return exigivel - mapeados

    linha_bp(
        "bp_caixa",
        "Caixa e equivalentes",
        "caixa_equivalentes",
        "caixa_equivalentes",
        lambda t: (f"={rp('dfc_caixa_eop', t)}", p("dfc", "caixa_final", t)),
    )
    linha_bp(
        "bp_aplicacoes",
        "Aplicacoes financeiras",
        "aplicacoes_financeiras",
        "aplicacoes_financeiras",
        f_constante(
            "bp_aplicacoes", "aplicacoes_financeiras", "aplicacoes_financeiras"
        ),
    )
    linha_bp(
        "bp_contas_receber",
        "Contas a receber",
        "contas_receber",
        "contas_receber",
        lambda t: (
            (f"={rp('wk_contas_receber', t)}", p("wk", "contas_receber", t))
            if "wk_contas_receber" in aba.linhas
            else None
        ),
    )
    linha_bp(
        "bp_estoques",
        "Estoques",
        "estoques",
        "estoques",
        lambda t: (
            (f"={rp('wk_estoques', t)}", p("wk", "estoques", t))
            if "wk_estoques" in aba.linhas
            else None
        ),
    )
    linha_bp(
        "bp_tributos",
        "Tributos a recuperar",
        "tributos_a_recuperar",
        "tributos_a_recuperar",
        (
            (
                lambda t: (
                    f"={rp('wk_tributos', t)}",
                    abs(p("wk", "tributos_a_recuperar", t) or 0.0),
                )
            )
            if "wk_tributos" in aba.linhas
            else None
        ),
    )
    linha_bp(
        "bp_imobilizado",
        "Imobilizado",
        "imobilizado",
        "imobilizado",
        lambda t: (f"={rp('ppe_eop', t)}", p("ppe", "imobilizado", t)),
    )
    linha_bp(
        "bp_intangivel",
        "Intangivel",
        "intangivel",
        "intangivel",
        f_constante("bp_intangivel", "intangivel", "intangivel"),
    )
    linha_bp(
        "bp_outros_ativos",
        "Outros ativos (residual do Ano 0)",
        "outros_ativos",
        None,
        f_constante("bp_outros_ativos", None, "outros_ativos"),
        hist_fn=hist_outros_ativos,
        comentario="Residual constante e explicito do BP REAL do Ano 0 "
        "(historico = Ativo Total CVM - linhas mapeadas).",
    )

    def f_ativo_total(t: int) -> tuple[str, float | None]:
        chaves = [
            "bp_caixa",
            "bp_aplicacoes",
            "bp_contas_receber",
            "bp_estoques",
            "bp_tributos",
            "bp_imobilizado",
            "bp_intangivel",
            "bp_outros_ativos",
        ]
        formula = "=" + "+".join(rp(c, t) for c in chaves)
        campos = [
            "caixa_equivalentes",
            "aplicacoes_financeiras",
            "contas_receber",
            "estoques",
            "tributos_a_recuperar",
            "imobilizado",
            "intangivel",
            "outros_ativos",
        ]
        valores = [b(c, t) for c in campos]
        soma = sum(x or 0.0 for x in valores)
        return formula, soma

    linha_bp(
        "bp_ativo_total",
        "(=) ATIVO TOTAL",
        "ativo_total",
        "ativo_total",
        f_ativo_total,
        negrito=True,
    )

    linha_bp(
        "bp_fornecedores",
        "Fornecedores",
        "fornecedores",
        "fornecedores",
        (
            (
                lambda t: (
                    f"={rp('wk_fornecedores', t)}",
                    abs(p("wk", "fornecedores", t) or 0.0),
                )
            )
            if "wk_fornecedores" in aba.linhas
            else None
        ),
    )
    linha_bp(
        "bp_obrigacoes",
        "Obrigacoes sociais e trabalhistas",
        "obrigacoes_sociais_trabalhistas",
        "obrigacoes_sociais_trabalhistas",
        (
            (
                lambda t: (
                    f"={rp('wk_obrigacoes', t)}",
                    abs(p("wk", "obrigacoes_sociais_trabalhistas", t) or 0.0),
                )
            )
            if "wk_obrigacoes" in aba.linhas
            else None
        ),
    )
    linha_bp(
        "bp_adiantamentos",
        "Adiantamento de clientes",
        "adiantamento_clientes",
        None,
        (
            (
                lambda t: (
                    f"={rp('wk_adiantamentos', t)}",
                    abs(p("wk", "adiantamento_clientes", t) or 0.0),
                )
            )
            if "wk_adiantamentos" in aba.linhas
            else None
        ),
    )
    linha_bp(
        "bp_divida_cp",
        "Emprestimos e financiamentos (CP)",
        "divida_curto_prazo",
        "divida_curto_prazo",
        lambda t: (f"={rp('div_cp', t)}", p("divida", "divida_curto_prazo", t)),
    )
    linha_bp(
        "bp_divida_lp",
        "Emprestimos e financiamentos (LP)",
        "divida_longo_prazo",
        "divida_longo_prazo",
        lambda t: (f"={rp('div_lp', t)}", p("divida", "divida_longo_prazo", t)),
    )
    linha_bp(
        "bp_arrendamento",
        "Passivo de arrendamento (IFRS-16)",
        "passivo_arrendamento",
        None,
        f_constante("bp_arrendamento", None, "passivo_arrendamento"),
    )
    linha_bp(
        "bp_outros_passivos",
        "Outros passivos (residual do Ano 0)",
        "outros_passivos",
        None,
        f_constante("bp_outros_passivos", None, "outros_passivos"),
        hist_fn=hist_outros_passivos,
        comentario="Residual constante do BP REAL do Ano 0.",
    )

    def f_passivo_total(t: int) -> tuple[str, float | None]:
        chaves = [
            "bp_fornecedores",
            "bp_obrigacoes",
            "bp_adiantamentos",
            "bp_divida_cp",
            "bp_divida_lp",
            "bp_arrendamento",
            "bp_outros_passivos",
        ]
        formula = "=" + "+".join(rp(c, t) for c in chaves)
        campos = [
            "fornecedores",
            "obrigacoes_sociais_trabalhistas",
            "adiantamento_clientes",
            "divida_curto_prazo",
            "divida_longo_prazo",
            "passivo_arrendamento",
            "outros_passivos",
        ]
        soma = sum(b(c, t) or 0.0 for c in campos)
        return formula, soma

    linha_bp(
        "bp_passivo_total",
        "(=) Passivo exigivel",
        "passivo_total",
        None,
        f_passivo_total,
        negrito=True,
        hist_fn=lambda i: (
            h("passivo_total", i) - h("patrimonio_liquido", i)
            if h("passivo_total", i) is not None
            and h("patrimonio_liquido", i) is not None
            else None
        ),
    )

    def f_pl(t: int) -> tuple[str, float | None]:
        ll = v("lucro_liquido", t)
        dividendos = p("dfc", "dividendos", t)
        if t == 1:
            base = h("patrimonio_liquido", nh - 1)
            ref_base = _ref("", CH(nh - 1), aba.L("bp_pl"))
        else:
            base = b("patrimonio_liquido", t - 1)
            ref_base = rp("bp_pl", t - 1)
        formula = f"={ref_base}+{rp('lucro_liquido', t)}+{rp('dfc_dividendos', t)}"
        valor = (
            base + ll - dividendos
            if base is not None and ll is not None and dividendos is not None
            else None
        )
        return formula, valor

    linha_bp(
        "bp_pl",
        "PATRIMONIO LIQUIDO",
        "patrimonio_liquido",
        "patrimonio_liquido",
        f_pl,
        negrito=True,
    )

    def f_passivo_pl(t: int) -> tuple[str, float | None]:
        return (
            f"={rp('bp_passivo_total', t)}+{rp('bp_pl', t)}",
            (b("passivo_total", t) or 0.0) + (b("patrimonio_liquido", t) or 0.0),
        )

    linha_bp(
        "bp_passivo_pl",
        "(=) PASSIVO + PL",
        "passivo_patrimonio_liquido",
        "passivo_total",
        f_passivo_pl,
        negrito=True,
    )

    # Linha Check (Direcional L122): booleana visivel.
    linha = aba.L("bp_check")
    aba.rotulo(linha, "CHECK (Ativo = Passivo + PL)", negrito=True)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        formula = (
            f'=IF(ROUND({rp("bp_ativo_total", t)},0)='
            f'ROUND({rp("bp_passivo_pl", t)},0),"Ok",'
            f'{rp("bp_ativo_total", t)}-{rp("bp_passivo_pl", t)})'
        )
        celula = ws.cell(row=linha, column=CP(t), value=formula)
        diferenca = b("diferenca_balanco", t) or 0.0
        ok = abs(diferenca) < 1.0
        celula.font = Font(
            name=FONTE_NUMERO,
            size=10,
            bold=True,
            color=COR_VERDE_FUNDO if ok else COR_VERMELHO_FUNDO,
        )
        celula.alignment = Alignment(horizontal="center")
        aba.valores[(linha, CP(t))] = "Ok" if ok else diferenca

    # ---- DFC indireto ----
    aba.titulo(aba.L("sec_dfc"), "DFC INDIRETO PROJETADO", col_p0 + 8)

    def d(campo: str, t: int) -> float | None:
        return p("dfc", campo, t)

    def linha_dfc(
        chave: str,
        rotulo: str,
        campo: str,
        formula_fn: Callable[[int], tuple[str, float | None]] | None,
        negrito: bool = False,
        serie_hist: str | None = None,
        sinal_motor: float = 1.0,
    ) -> None:
        linha = aba.L(chave)
        aba.rotulo(linha, rotulo, negrito=negrito, recuo=1)
        for i in range(nh):
            valor_hist = h(serie_hist, i) if serie_hist else None
            aba.numero(linha, CH(i), valor_hist, FORMATO_MILHAR, cor=COR_HISTORICO)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            bruto = d(campo, t)
            valor_motor = bruto * sinal_motor if bruto is not None else None
            if formula_fn is None or valor_motor is None:
                aba.numero(linha, CP(t), valor_motor, FORMATO_MILHAR, negrito=negrito)
                continue
            formula, valor_python = formula_fn(t)
            aba.calculo(
                linha,
                CP(t),
                formula,
                valor_python,
                valor_motor,
                FORMATO_MILHAR,
                negrito=negrito,
            )

    linha_dfc(
        "dfc_ll",
        "Lucro liquido",
        "lucro_liquido",
        lambda t: (f"={rp('lucro_liquido', t)}", v("lucro_liquido", t)),
        serie_hist="lucro_liquido",
    )
    linha_dfc(
        "dfc_da",
        "(+) Depreciacao e amortizacao",
        "depreciacao_amortizacao",
        lambda t: (f"=-{rp('dre_da', t)}", v("depreciacao_amortizacao", t)),
        serie_hist="depreciacao_amortizacao",
    )

    campo_var = {
        "wk_contas_receber": ("variacao_contas_receber", -1.0),
        "wk_estoques": ("variacao_estoques", -1.0),
        "wk_tributos": ("variacao_tributos_a_recuperar", -1.0),
        "wk_fornecedores": ("variacao_fornecedores", 1.0),
        "wk_obrigacoes": ("variacao_obrigacoes_sociais_trabalhistas", 1.0),
        "wk_adiantamentos": ("variacao_adiantamento_clientes", 1.0),
    }
    for chave_wk in ctx["_contas_wk_presentes"]:
        campo, sentido = campo_var[chave_wk]
        rotulo_conta = CONTAS_WK_INFO[chave_wk][0]

        def f_var(
            t: int, chave_wk=chave_wk, sentido=sentido
        ) -> tuple[str, float | None]:
            campo_wk = CONTAS_WK_INFO[chave_wk][1]
            atual = abs(p("wk", campo_wk, t) or 0.0)
            if t == 1:
                anterior = abs(
                    _numero(ctx["projecao"].get("ano0", {}).get("wk", {}).get(campo_wk))
                    or 0.0
                )
                serie_nome = CONTAS_WK_INFO[chave_wk][2]
                if serie_nome and h(serie_nome, nh - 1) is not None:
                    ref_ant = _ref("", CH(nh - 1), aba.L(chave_wk))
                    anterior_ref = abs(h(serie_nome, nh - 1))
                else:
                    return "", None  # sem base historica -> valor colado
                anterior = anterior_ref
            else:
                anterior = abs(p("wk", campo_wk, t - 1) or 0.0)
                ref_ant = rp(chave_wk, t - 1)
            sinal = "-" if sentido < 0 else ""
            formula = f"={sinal}({rp(chave_wk, t)}-{ref_ant})"
            return formula, sentido * (atual - anterior)

        linha_dfc(
            f"dfc_var_{chave_wk}", f"(+/-) Variacao de {rotulo_conta}", campo, f_var
        )

    def f_fco(t: int) -> tuple[str, float | None]:
        chaves = ["dfc_ll", "dfc_da"] + [
            f"dfc_var_{c}" for c in ctx["_contas_wk_presentes"]
        ]
        formula = "=" + "+".join(rp(c, t) for c in chaves)
        ll, da = d("lucro_liquido", t), d("depreciacao_amortizacao", t)
        delta = d("delta_nwc", t)
        valor = (
            ll + da - delta
            if ll is not None and da is not None and delta is not None
            else None
        )
        return formula, valor

    linha_dfc(
        "dfc_fco",
        "(=) Caixa operacional (FCO)",
        "fco",
        f_fco,
        negrito=True,
        serie_hist="fco",
    )
    linha_dfc(
        "dfc_capex",
        "(-) CAPEX",
        "capex",
        lambda t: (f"={rp('ppe_capex', t)}", d("capex", t)),
    )
    linha_dfc(
        "dfc_fci",
        "(=) Caixa de investimento (FCI)",
        "fci",
        lambda t: (f"={rp('dfc_capex', t)}", d("capex", t)),
        negrito=True,
        serie_hist="fci",
    )
    linha_dfc(
        "dfc_dividendos",
        "(-) Dividendos pagos",
        "dividendos",
        lambda t: (
            f"=-MAX({rp('lucro_liquido', t)},0)"
            f"*{prem('payout_dividendos', 0, travar=True)}",
            -(
                max(v("lucro_liquido", t) or 0.0, 0.0)
                * (_numero(premissas.get("payout_dividendos")) or 0.0)
            ),
        ),
        sinal_motor=-1.0,
    )
    linha_dfc(
        "dfc_delta_divida",
        "(+/-) Variacao de divida",
        "delta_divida",
        lambda t: (
            f"={rp('div_eop', t)}-{rp('div_bop', t)}",
            (p("divida", "divida_bruta", t) or 0.0)
            - (p("divida", "divida_abertura", t) or 0.0),
        ),
    )

    def f_fcfin(t: int) -> tuple[str, float | None]:
        formula = f"={rp('dfc_delta_divida', t)}+{rp('dfc_dividendos', t)}"
        delta, dividendos = d("delta_divida", t), d("dividendos", t)
        valor = (
            delta - dividendos if delta is not None and dividendos is not None else None
        )
        return formula, valor

    linha_dfc(
        "dfc_fcfin",
        "(=) Caixa de financiamento (FCFin)",
        "fcfin",
        f_fcfin,
        negrito=True,
        serie_hist="fcf",
    )

    def f_caixa_bop(t: int) -> tuple[str, float | None]:
        if t == 1:
            ref_base = _ref("", CH(nh - 1), aba.L("bp_caixa"))
            return f"={ref_base}", h("caixa_equivalentes", nh - 1)
        return f"={rp('dfc_caixa_eop', t - 1)}", d("caixa_final", t - 1)

    linha_dfc(
        "dfc_caixa_bop",
        "Caixa no inicio do periodo",
        "caixa_inicial",
        f_caixa_bop,
        serie_hist="caixa_inicial_dfc",
    )

    def f_variacao(t: int) -> tuple[str, float | None]:
        formula = f"={rp('dfc_fco', t)}+{rp('dfc_fci', t)}+{rp('dfc_fcfin', t)}"
        fco, fci, fcfin = d("fco", t), d("fci", t), d("fcfin", t)
        valor = (
            fco + fci + fcfin
            if fco is not None and fci is not None and fcfin is not None
            else None
        )
        return formula, valor

    linha_dfc(
        "dfc_variacao",
        "Variacao de caixa no periodo",
        "variacao_caixa",
        f_variacao,
        serie_hist="variacao_caixa",
    )
    linha_dfc(
        "dfc_caixa_eop",
        "Caixa no fim do periodo (= BP)",
        "caixa_final",
        lambda t: (
            f"={rp('dfc_caixa_bop', t)}+{rp('dfc_variacao', t)}",
            (d("caixa_inicial", t) or 0.0) + (d("variacao_caixa", t) or 0.0),
        ),
        negrito=True,
        serie_hist="caixa_final_dfc",
    )

    # ---- WK multi-driver ----
    aba.titulo(
        aba.L("sec_wk"), "CAPITAL DE GIRO (conta = dias x driver / 365)", col_p0 + 8
    )
    driver_ref = {
        "receita_liquida": "receita_liquida",
        "cpv": "cpv",
        "sgna": "sgna",
    }
    for chave_wk in ctx["_contas_wk_presentes"]:
        rotulo_conta, campo_wk, serie_nome, driver, rotulo_dias = CONTAS_WK_INFO[
            chave_wk
        ]
        linha = aba.L(chave_wk)
        aba.rotulo(linha, rotulo_conta, recuo=1)
        for i in range(nh):
            valor_hist = (
                abs(h(serie_nome, i))
                if serie_nome and h(serie_nome, i) is not None
                else None
            )
            aba.numero(linha, CH(i), valor_hist, FORMATO_MILHAR, cor=COR_HISTORICO)

        # Dias IMPLICITOS do motor (constante nos 8 anos): celula VERDE
        # editavel; saldo = dias x driver / 365.
        saldo_1 = abs(p("wk", campo_wk, 1) or 0.0)
        driver_1 = abs(v(driver_ref.get(driver, "receita_liquida"), 1) or 0.0)
        dias = saldo_1 * 365.0 / driver_1 if driver_1 else None
        linha_dias = aba.L(f"{chave_wk}_dias")
        aba.rotulo(linha_dias, rotulo_dias, recuo=2)
        aba.numero(
            linha_dias,
            CP(1),
            dias,
            FORMATO_DIAS,
            cor=COR_PREMISSA,
            comentario="Dias implicitos do motor (edite para simular).",
        )
        chave_driver = {
            "receita_liquida": "receita_liquida",
            "cpv": "cpv",
            "sgna": "sgna",
        }[driver if driver in driver_ref else "receita_liquida"]
        for t in range(1, HORIZONTE_PROJECAO + 1):
            valor_motor = abs(p("wk", campo_wk, t) or 0.0)
            driver_t = abs(v(chave_driver, t) or 0.0)
            valor_python = (
                (dias or 0.0) * driver_t / 365.0 if dias is not None else None
            )
            ref_dias = _ref("", CP(1), linha_dias, travar=True)
            sinal_driver = "-" if chave_driver in ("cpv", "sgna") else ""
            formula = f"={ref_dias}*{sinal_driver}{rp(chave_driver, t)}/365"
            aba.calculo(
                linha, CP(t), formula, valor_python, valor_motor, FORMATO_MILHAR
            )

    def f_nwc(t: int) -> tuple[str, float | None]:
        ativos = [
            c
            for c in ("wk_contas_receber", "wk_estoques", "wk_tributos")
            if c in aba.linhas
            and f"{c}_dias" in aba.linhas
            and c in ctx["_contas_wk_presentes"]
        ]
        passivos = [
            c
            for c in ("wk_fornecedores", "wk_obrigacoes", "wk_adiantamentos")
            if c in ctx["_contas_wk_presentes"]
        ]
        termos = [f"+{rp(c, t)}" for c in ativos] + [f"-{rp(c, t)}" for c in passivos]
        formula = "=" + "".join(termos).lstrip("+")
        soma = sum(abs(p("wk", CONTAS_WK_INFO[c][1], t) or 0.0) for c in ativos)
        soma -= sum(abs(p("wk", CONTAS_WK_INFO[c][1], t) or 0.0) for c in passivos)
        return formula, soma

    linha = aba.L("wk_nwc")
    aba.rotulo(linha, "(=) NWC (ativos - passivos de giro)", negrito=True, recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        campo_nwc = (
            "nwc_multi_driver"
            if _proj(ctx, "wk", "nwc_multi_driver", t) is not None
            else "nwc"
        )
        formula, valor_python = f_nwc(t)
        aba.calculo(
            linha,
            CP(t),
            formula,
            valor_python,
            p("wk", campo_nwc, t),
            FORMATO_MILHAR,
            negrito=True,
        )

    linha = aba.L("wk_delta_nwc")
    aba.rotulo(linha, "Δ NWC (positivo = consome caixa)", recuo=2)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        valor_motor = p("wk", "delta_nwc", t)
        if t == 1:
            aba.numero(linha, CP(t), valor_motor, FORMATO_MILHAR)
            continue
        campo_nwc = (
            "nwc_multi_driver"
            if _proj(ctx, "wk", "nwc_multi_driver", t) is not None
            else "nwc"
        )
        atual, anterior = p("wk", campo_nwc, t), p("wk", campo_nwc, t - 1)
        valor_python = (
            atual - anterior if atual is not None and anterior is not None else None
        )
        aba.calculo(
            linha,
            CP(t),
            f"={rp('wk_nwc', t)}-{rp('wk_nwc', t - 1)}",
            valor_python,
            valor_motor,
            FORMATO_MILHAR,
        )

    # ---- Divida ----
    aba.titulo(
        aba.L("sec_divida"), "DIVIDA (rolagem + captacao automatica)", col_p0 + 8
    )

    def linha_div(
        chave: str,
        rotulo: str,
        campo: str,
        formula_fn: Callable[[int], tuple[str, float | None]] | None = None,
        formato: str = FORMATO_MILHAR,
        cor: str = COR_FORMULA,
        negrito: bool = False,
        sinal: float = 1.0,
    ) -> None:
        linha = aba.L(chave)
        aba.rotulo(linha, rotulo, negrito=negrito, recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            bruto = p("divida", campo, t)
            valor_motor = bruto * sinal if bruto is not None else None
            if formula_fn is None or valor_motor is None:
                aba.numero(linha, CP(t), valor_motor, formato, cor=cor, negrito=negrito)
                continue
            formula, valor_python = formula_fn(t)
            aba.calculo(
                linha,
                CP(t),
                formula,
                valor_python,
                valor_motor,
                formato,
                cor=cor,
                negrito=negrito,
            )

    def f_div_bop(t: int) -> tuple[str, float | None]:
        if t == 1:
            return "", None  # valor do ano 0 colado
        return f"={rp('div_eop', t - 1)}", p("divida", "divida_bruta", t - 1)

    linha_div("div_bop", "Divida BoP", "divida_abertura", f_div_bop)
    linha_div("div_captacao", "(+) Captacao (caixa minimo)", "captacao")
    linha_div("div_amortizacao", "(-) Amortizacao", "amortizacao", sinal=-1.0)

    def f_div_eop(t: int) -> tuple[str, float | None]:
        formula = (
            f"={rp('div_bop', t)}+{rp('div_captacao', t)}+{rp('div_amortizacao', t)}"
        )
        bop = p("divida", "divida_abertura", t)
        capt = p("divida", "captacao", t)
        amort = p("divida", "amortizacao", t)
        valor = (
            bop + capt - amort
            if bop is not None and capt is not None and amort is not None
            else None
        )
        return formula, valor

    linha_div("div_eop", "Divida EoP", "divida_bruta", f_div_eop, negrito=True)
    linha_div("div_cp", "Divida de curto prazo (reclassificada)", "divida_curto_prazo")
    linha_div("div_lp", "Divida de longo prazo", "divida_longo_prazo")
    linha_div(
        "div_kd",
        "Custo da divida (Kd)",
        "custo_divida_kd",
        formato=FORMATO_PERCENTUAL_2,
        cor=COR_PREMISSA,
    )

    def f_juros(t: int) -> tuple[str, float | None]:
        bop = p("divida", "divida_abertura", t)
        kd = p("divida", "custo_divida_kd", t)
        formula = f"=-{rp('div_bop', t)}*{rp('div_kd', t)}"
        return formula, (-bop * kd if bop is not None and kd is not None else None)

    linha_div(
        "div_juros", "(-) Juros da divida (Kd x BoP)", "juros", f_juros, sinal=-1.0
    )
    linha_div(
        "div_taxa_aplicacao",
        "Taxa de aplicacao do caixa (CDI)",
        "taxa_aplicacao_caixa",
        formato=FORMATO_PERCENTUAL_2,
    )

    def f_recfin(t: int) -> tuple[str, float | None]:
        taxa = p("divida", "taxa_aplicacao_caixa", t)
        if t == 1:
            caixa_ant = h("caixa_equivalentes", nh - 1)
            ref_caixa = _ref("", CH(nh - 1), aba.L("bp_caixa"))
        else:
            caixa_ant = p("dfc", "caixa_final", t - 1)
            ref_caixa = rp("dfc_caixa_eop", t - 1)
        aplic = p("balanco", "aplicacoes_financeiras", t)
        formula = (
            f"={rp('div_taxa_aplicacao', t)}"
            f"*MAX({ref_caixa}+{rp('bp_aplicacoes', t)},0)"
        )
        valor = (
            taxa * max((caixa_ant or 0.0) + (aplic or 0.0), 0.0)
            if taxa is not None and caixa_ant is not None
            else None
        )
        return formula, valor

    linha_div(
        "div_receita_financeira",
        "(+) Receita financeira s/ caixa",
        "receita_financeira_caixa",
        f_recfin,
    )

    # ---- Leasing (condicional) ----
    if ctx["_tem_leasing"]:
        aba.titulo(aba.L("sec_leasing"), "ARRENDAMENTO (IFRS-16)", col_p0 + 8)
        linha = aba.L("leasing_passivo")
        aba.rotulo(linha, "Passivo de arrendamento (constante Ano 0)", recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            aba.numero(linha, CP(t), p("balanco", "passivo_arrendamento", t))
        linha = aba.L("leasing_juros")
        aba.rotulo(linha, "(-) Juros de arrendamento", recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            juros_arr = p("divida", "juros_arrendamento", t)
            aba.numero(linha, CP(t), -juros_arr if juros_arr is not None else None)
        linha = aba.L("leasing_da_direito_uso")
        aba.rotulo(linha, "D&A do direito de uso (reclassificada)", recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            aba.numero(linha, CP(t), p("dre", "da_direito_uso", t))

    # ---- PP&E ----
    aba.titulo(
        aba.L("sec_ppe"), "PP&E (CAPEX = % receita; D&A = % do PP&E BoP)", col_p0 + 8
    )

    def f_ppe_bop(t: int) -> tuple[str, float | None]:
        if t == 1:
            ref_base = _ref("", CH(nh - 1), aba.L("bp_imobilizado"))
            return f"={ref_base}", h("imobilizado", nh - 1)
        return f"={rp('ppe_eop', t - 1)}", p("ppe", "imobilizado", t - 1)

    linha = aba.L("ppe_bop")
    aba.rotulo(linha, "PP&E BoP (inicio do periodo)", recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        bop_motor = (
            _numero(ctx["projecao"].get("ano0", {}).get("ppe", {}).get("imobilizado"))
            if t == 1
            else p("ppe", "imobilizado", t - 1)
        )
        formula, valor_python = f_ppe_bop(t)
        aba.calculo(linha, CP(t), formula, valor_python, bop_motor, FORMATO_MILHAR)

    linha = aba.L("ppe_capex_pct")
    aba.rotulo(linha, "Capex % da receita (premissa)", recuo=2)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        pct = _numero(premissas.get(f"capex_receita_ano{t}"))
        aba.calculo(
            linha,
            CP(t),
            f"={prem('capex_receita', t)}",
            pct,
            pct,
            FORMATO_PERCENTUAL_2,
            cor=COR_PREMISSA,
        )

    def f_ppe_capex(t: int) -> tuple[str, float | None]:
        rl = v("receita_liquida", t)
        pct = _numero(premissas.get(f"capex_receita_ano{t}"))
        formula = f"={rp('ppe_capex_pct', t)}*{rp('receita_liquida', t)}"
        return formula, (pct * rl if pct is not None and rl is not None else None)

    linha = aba.L("ppe_capex")
    aba.rotulo(linha, "(+) CAPEX (negativo = saida de caixa)", recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        formula, valor_python = f_ppe_capex(t)
        aba.calculo(
            linha, CP(t), formula, valor_python, p("ppe", "capex", t), FORMATO_MILHAR
        )

    politica_ppe = ctx["projecao"].get("politicas_projecao", {}).get("ppe", {})
    taxa_da = _numero(politica_ppe.get("taxa_depreciacao_anual"))
    if taxa_da is None:
        da_1 = p("ppe", "da_imobilizado", 1) or p("ppe", "depreciacao_amortizacao", 1)
        bop_1 = _numero(
            ctx["projecao"].get("ano0", {}).get("ppe", {}).get("imobilizado")
        )
        taxa_da = (da_1 / bop_1) if da_1 and bop_1 else None
    linha = aba.L("ppe_da_pct")
    aba.rotulo(
        linha,
        "D&A % do PP&E (usada — config/premissa)",
        recuo=2,
        comentario="Taxa usada pelo motor (1/vida util). A media historica e "
        "so INFORMACAO (linha abaixo) — decisao D-047.",
    )
    aba.numero(linha, CP(1), taxa_da, FORMATO_PERCENTUAL_2, cor=COR_PREMISSA)

    def f_ppe_da(t: int) -> tuple[str, float | None]:
        bop = (
            _numero(ctx["projecao"].get("ano0", {}).get("ppe", {}).get("imobilizado"))
            if t == 1
            else p("ppe", "imobilizado", t - 1)
        )
        ref_taxa = _ref("", CP(1), aba.L("ppe_da_pct"), travar=True)
        formula = f"=-MIN({ref_taxa}*{rp('ppe_bop', t)},{rp('ppe_bop', t)})"
        valor = (
            -min((taxa_da or 0.0) * bop, bop) if bop is not None and taxa_da else None
        )
        return formula, valor

    linha = aba.L("ppe_da")
    aba.rotulo(linha, "(-) D&A do periodo", recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        da_motor = p("dre", "depreciacao_amortizacao", t)
        formula, valor_python = f_ppe_da(t)
        aba.calculo(
            linha,
            CP(t),
            formula,
            valor_python,
            -da_motor if da_motor is not None else None,
            FORMATO_MILHAR,
        )

    def f_ppe_eop(t: int) -> tuple[str, float | None]:
        bop = (
            _numero(ctx["projecao"].get("ano0", {}).get("ppe", {}).get("imobilizado"))
            if t == 1
            else p("ppe", "imobilizado", t - 1)
        )
        capex = p("ppe", "capex", t)
        da = p("dre", "depreciacao_amortizacao", t)
        formula = f"={rp('ppe_bop', t)}-{rp('ppe_capex', t)}+{rp('ppe_da', t)}"
        valor = (
            bop + abs(capex) - da
            if bop is not None and capex is not None and da is not None
            else None
        )
        return formula, valor

    linha = aba.L("ppe_eop")
    aba.rotulo(linha, "PP&E EoP (final do periodo)", negrito=True, recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        formula, valor_python = f_ppe_eop(t)
        aba.calculo(
            linha,
            CP(t),
            formula,
            valor_python,
            p("ppe", "imobilizado", t),
            FORMATO_MILHAR,
            negrito=True,
        )

    linha = aba.L("ppe_da_pct_historica")
    aba.rotulo(linha, "D&A % do PP&E — media historica (INFO)", recuo=2)
    da_hist = _numero(
        ctx["projecao"].get("ano0", {}).get("ppe", {}).get("da_pct_ppe_historica")
    )
    aba.numero(linha, CP(1), da_hist, FORMATO_PERCENTUAL_2, cor=COR_HISTORICO)

    linha = aba.L("ppe_intangivel")
    aba.rotulo(linha, "Intangivel (constante, linha propria do BP)", recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        aba.numero(linha, CP(t), p("ppe", "intangivel", t))

    return aba


# ---------------------------------------------------------------------------
# Abas 4 e 5 — FCFF e FCFE (SEPARADAS, referenciando a aba Modelo)
# ---------------------------------------------------------------------------


def _cp_modelo(ctx: dict[str, Any], t: int) -> int:
    """Coluna projetada do ano t na aba Modelo."""
    return 2 + len(ctx["anos_historicos"]) + (t - 1)


def _rm(ctx: dict[str, Any], aba_modelo: Aba, chave: str, t: int) -> str:
    """Referencia 'Modelo!<coluna do ano t><linha da chave>'."""
    return _ref("Modelo", _cp_modelo(ctx, t), aba_modelo.L(chave))


def _aba_fcff(wb: Workbook, ctx: dict[str, Any], aba_modelo: Aba) -> Aba:
    """FCFF: NOPAT -> FCFF -> VP -> EV -> Equity -> Target; WACC; retornos."""
    aba = Aba(wb, "FCFF")
    ws = aba.ws
    ws.column_dimensions["A"].width = 40
    for coluna in range(2, 11):
        ws.column_dimensions[get_column_letter(coluna)].width = 14

    projecao = ctx["projecao"]
    fcff = projecao["fcff"]
    wacc_bloco = projecao.get("wacc", {})
    vt = projecao.get("valor_terminal", {})
    ev_equity = projecao.get("ev_equity", {})
    retornos = projecao.get("retornos", {})

    def fx(campo: str, t: int) -> float | None:
        return _numero(fcff.get(f"ano{t}", {}).get(campo))

    aba.titulo(1, f"FCFF — DCF pela firma ({ctx['ticker']})", 10)
    aba.cabecalho(2, 2, [f"Ano {t}" for t in range(1, HORIZONTE_PROJECAO + 1)])

    aliquota_nopat = _numero(fcff.get("ano1", {}).get("aliquota_ir_nopat")) or 0.0
    wacc_usado = _numero(wacc_bloco.get("wacc"))
    g = _numero(vt.get("g"))
    conv = ctx["convencao_desconto"]

    def _exp(ano: int) -> float:
        """Expoente de desconto do ano (mesma convencao do motor)."""
        return expoente_desconto(ano, conv)

    def _exp_txt(ano: int) -> str:
        """Expoente formatado para a formula (8, nao 8.0, em periodo cheio)."""
        return f"{_exp(ano):g}"

    # Layout vertical com registro de linhas.
    cursor = 3

    def linha_serie(
        chave: str,
        rotulo: str,
        formula_fn: Callable[[int], tuple[str, float | None]] | None,
        valores_motor: Callable[[int], float | None],
        formato: str = FORMATO_MILHAR,
        negrito: bool = False,
    ) -> None:
        nonlocal cursor
        aba.registrar(chave, cursor)
        aba.rotulo(cursor, rotulo, negrito=negrito, recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            valor_motor = valores_motor(t)
            if formula_fn is None or valor_motor is None:
                aba.numero(cursor, 1 + t, valor_motor, formato, negrito=negrito)
                continue
            formula, valor_python = formula_fn(t)
            aba.calculo(
                cursor,
                1 + t,
                formula,
                valor_python,
                valor_motor,
                formato,
                negrito=negrito,
            )
        cursor += 1

    def rl(chave: str, t: int, travar: bool = False) -> str:
        return _ref("", 1 + t, aba.L(chave), travar)

    # A celula do WACC usado fica no build-up (registrada depois); a formula
    # de VP usa um NOME DEFINIDO 'WACC' para legibilidade e Data Tables.
    linha_serie(
        "ebit",
        "EBIT (Modelo)",
        lambda t: (f"={_rm(ctx, aba_modelo, 'ebit', t)}", _proj(ctx, "dre", "ebit", t)),
        lambda t: _proj(ctx, "dre", "ebit", t),
    )
    linha_serie(
        "nopat",
        "NOPAT = EBIT x (1 - T)",
        lambda t: (
            f"={rl('ebit', t)}*(1-{aliquota_nopat})",
            (_proj(ctx, "dre", "ebit", t) or 0.0) * (1 - aliquota_nopat),
        ),
        lambda t: fx("nopat", t),
        negrito=True,
    )
    linha_serie(
        "da",
        "(+) Depreciacao e amortizacao",
        lambda t: (
            f"=-{_rm(ctx, aba_modelo, 'dre_da', t)}",
            fx("depreciacao_amortizacao", t),
        ),
        lambda t: fx("depreciacao_amortizacao", t),
    )
    linha_serie(
        "delta_wk",
        "(-) Δ Working Capital",
        lambda t: (
            f"=-{_rm(ctx, aba_modelo, 'wk_delta_nwc', t)}",
            -(fx("delta_nwc", t) or 0.0),
        ),
        lambda t: (
            -(fx("delta_nwc", t) or 0.0) if fx("delta_nwc", t) is not None else None
        ),
    )
    linha_serie(
        "capex",
        "(-) CAPEX",
        lambda t: (
            f"={_rm(ctx, aba_modelo, 'ppe_capex', t)}",
            _proj(ctx, "ppe", "capex", t),
        ),
        lambda t: (
            -(fx("capex_saida_caixa", t) or 0.0)
            if fx("capex_saida_caixa", t) is not None
            else None
        ),
    )

    def f_fcff(t: int) -> tuple[str, float | None]:
        formula = (
            f"={rl('nopat', t)}+{rl('da', t)}+{rl('delta_wk', t)}+{rl('capex', t)}"
        )
        valores = [
            fx("nopat", t),
            fx("depreciacao_amortizacao", t),
            -(fx("delta_nwc", t) or 0.0),
            -(fx("capex_saida_caixa", t) or 0.0),
        ]
        valor = (
            sum(x for x in valores if x is not None) if valores[0] is not None else None
        )
        return formula, valor

    linha_serie("fcff", "(=) FCFF", f_fcff, lambda t: fx("fcff", t), negrito=True)
    # t (anos) guarda o EXPOENTE de desconto (t, ou t-0,5 em meio-periodo) para
    # a formula de VP referenciar — igual a convencao do motor.
    linha_serie("t_anos", "t (desconto)", None, lambda t: _exp(t), FORMATO_MILHAR_2)

    def f_vp(t: int) -> tuple[str, float | None]:
        fluxo = fx("fcff", t)
        valor = (
            fluxo / (1 + wacc_usado) ** _exp(t)
            if fluxo is not None and wacc_usado
            else None
        )
        return f"={rl('fcff', t)}/(1+WACC)^{rl('t_anos', t)}", valor

    linha_serie(
        "vp_fcff",
        "VP do FCFF",
        f_vp,
        lambda t: (
            (fx("fcff", t) or 0.0) / (1 + wacc_usado) ** _exp(t) if wacc_usado else None
        ),
    )
    linha_serie("roic", "ROIC (%)", None, lambda t: fx("roic", t), FORMATO_PERCENTUAL)
    linha_serie(
        "roiic", "ROIIC (%)", None, lambda t: fx("roiic", t), FORMATO_PERCENTUAL
    )

    cursor += 1
    aba.titulo(cursor, "ENTERPRISE VALUE -> EQUITY -> TARGET (bridge)", 10)
    cursor += 1

    def linha_valor(
        chave: str,
        rotulo: str,
        valor_motor: Any,
        formula: str | None = None,
        valor_python: Any = None,
        formato: str = FORMATO_MILHAR,
        cor: str = COR_FORMULA,
        negrito: bool = False,
    ) -> None:
        nonlocal cursor
        aba.registrar(chave, cursor)
        aba.rotulo(cursor, rotulo, negrito=negrito, recuo=1)
        if formula is not None:
            aba.calculo(
                cursor,
                2,
                formula,
                valor_python,
                valor_motor,
                formato,
                cor=cor,
                negrito=negrito,
            )
        else:
            aba.numero(cursor, 2, valor_motor, formato, cor=cor, negrito=negrito)
        cursor += 1

    soma_vp = _numero(vt.get("soma_vp_fcff")) or _numero(ev_equity.get("soma_vp_fcff"))
    soma_python = (
        sum(
            (fx("fcff", t) or 0.0) / (1 + wacc_usado) ** _exp(t)
            for t in range(1, HORIZONTE_PROJECAO + 1)
        )
        if wacc_usado
        else None
    )
    linha_valor(
        "soma_vp",
        "PV do fluxo explicito (anos 1-8)",
        soma_vp,
        formula=f"=SUM({_ref('', 2, aba.L('vp_fcff'))}:"
        f"{_ref('', 9, aba.L('vp_fcff'))})",
        valor_python=soma_python,
    )

    vt_bruto = _numero(vt.get("vt_bruto"))
    fcff8 = fx("fcff", 8)
    vt_python = (
        fcff8 * (1 + g) / (wacc_usado - g)
        if fcff8 is not None and wacc_usado and g is not None
        else None
    )
    linha_valor(
        "vt_bruto",
        f"Valor terminal (base: {vt.get('base_vt', 'fcff_ano8')})",
        vt_bruto,
        formula=f"={rl('fcff', 8)}*(1+g_perpetuidade)/(WACC-g_perpetuidade)",
        valor_python=vt_python,
    )
    vp_vt = _numero(vt.get("vp_vt"))
    linha_valor(
        "vp_vt",
        "PV da perpetuidade",
        vp_vt,
        formula=f"={_ref('', 2, aba.L('vt_bruto'))}/(1+WACC)^{_exp_txt(8)}",
        valor_python=(
            vt_bruto / (1 + wacc_usado) ** _exp(8)
            if vt_bruto is not None and wacc_usado
            else None
        ),
    )
    ev = _numero(ev_equity.get("ev"))
    linha_valor(
        "ev",
        "(=) Enterprise Value",
        ev,
        formula=f"={_ref('', 2, aba.L('soma_vp'))}+{_ref('', 2, aba.L('vp_vt'))}",
        valor_python=(soma_vp or 0.0) + (vp_vt or 0.0),
        negrito=True,
    )

    ajustes = ev_equity.get("ajustes_bridge", {})
    itens_bridge = (
        (
            "aj_divida",
            "(-) Divida bruta (CP+LP+leasing, Ano 0)",
            -(_numero(ajustes.get("divida_bruta")) or 0.0),
        ),
        (
            "aj_caixa",
            "(+) Caixa e equivalentes (Ano 0)",
            _numero(ajustes.get("caixa_equivalentes")),
        ),
        (
            "aj_aplicacoes",
            "(+) Aplicacoes financeiras (Ano 0)",
            _numero(ajustes.get("aplicacoes_financeiras")),
        ),
        (
            "aj_minoritarios",
            "(-) Participacoes minoritarias",
            -(_numero(ajustes.get("participacoes_minoritarias")) or 0.0),
        ),
        (
            "aj_coligadas",
            "(+) Investimentos em coligadas",
            _numero(ajustes.get("investimentos_coligadas")),
        ),
        (
            "aj_nao_op",
            "(+) Ativos nao operacionais",
            _numero(ajustes.get("ativos_nao_operacionais")),
        ),
    )
    for chave, rotulo, valor in itens_bridge:
        # Dados do BP REAL do Ano 0 => historicos, em AZUL.
        linha_valor(chave, rotulo, valor, cor=COR_HISTORICO)

    equity = _numero(ev_equity.get("equity_value"))
    formula_equity = "=" + "+".join(
        _ref("", 2, aba.L(c))
        for c in (
            "ev",
            "aj_divida",
            "aj_caixa",
            "aj_aplicacoes",
            "aj_minoritarios",
            "aj_coligadas",
            "aj_nao_op",
        )
    )
    equity_python = (ev or 0.0) + sum((v or 0.0) for _, _, v in itens_bridge)
    linha_valor(
        "equity",
        "(=) Equity Value",
        equity,
        formula=formula_equity,
        valor_python=equity_python,
        negrito=True,
    )

    acoes = _numero(ev_equity.get("acoes_fully_diluted"))
    fator = _numero(ev_equity.get("fator_escala_moeda")) or 1.0
    linha_valor("acoes", "Acoes fully diluted", acoes, formato="#,##0")
    linha_valor("fator", "Fator de escala (DFs em MIL)", fator, formato="#,##0")
    target = _numero(ev_equity.get("target_price"))
    linha_valor(
        "target",
        "TARGET PRICE (FCFF)",
        target,
        formula=f"={_ref('', 2, aba.L('equity'))}*{_ref('', 2, aba.L('fator'))}"
        f"/{_ref('', 2, aba.L('acoes'))}",
        valor_python=(equity * fator / acoes if equity is not None and acoes else None),
        formato=FORMATO_PRECO,
        negrito=True,
    )
    preco = _numero(ev_equity.get("preco_atual"))
    linha_valor("preco", "Preco atual", preco, formato=FORMATO_PRECO)
    linha_valor(
        "upside",
        "Upside",
        _numero(ev_equity.get("upside")),
        formula=f"={_ref('', 2, aba.L('target'))}/{_ref('', 2, aba.L('preco'))}-1",
        valor_python=(target / preco - 1 if target is not None and preco else None),
        formato=FORMATO_PERCENTUAL,
        negrito=True,
    )

    cursor += 1
    aba.titulo(cursor, "WACC — BUILD-UP (Ke CAPM Brasil + Kd)", 10)
    cursor += 1
    manual = wacc_bloco.get("wacc_origem") == "manual_do_analista"
    itens_wacc = (
        ("w_rf", "Risk-free USD (T-bond 10y)", wacc_bloco.get("rf_usd"), None, None),
        (
            "w_beta",
            "Beta (input Bloomberg)",
            wacc_bloco.get("beta_input", wacc_bloco.get("beta_desalavancado")),
            None,
            None,
        ),
        (
            "w_de",
            "D/E medio (anos 1-8, info)",
            wacc_bloco.get("divida_sobre_equity"),
            None,
            None,
        ),
        (
            "w_beta_l",
            "Beta usado (sem re-alavancagem)",
            wacc_bloco.get("beta_realavancado"),
            "=w_beta",
            None,
        ),
        ("w_erp", "ERP EUA", wacc_bloco.get("erp_eua"), None, None),
        ("w_crp", "CRP Brasil", wacc_bloco.get("crp_brasil"), None, None),
        ("w_ke_usd", "Ke USD", wacc_bloco.get("ke_usd"), None, None),
        ("w_ipca", "IPCA longo prazo", wacc_bloco.get("ipca"), None, None),
        ("w_cpi", "CPI EUA longo prazo", wacc_bloco.get("cpi_eua"), None, None),
        ("w_ke", "Ke BRL", wacc_bloco.get("ke_brl"), None, None),
        ("w_kd", "Kd (input CDI+spread)", wacc_bloco.get("kd_historico"), None, None),
        (
            "w_t",
            "Aliquota IR (escudo fiscal)",
            wacc_bloco.get("aliquota_ir"),
            None,
            None,
        ),
        ("w_pe", "Peso Equity (E/V)", wacc_bloco.get("peso_equity"), None, None),
        ("w_pd", "Peso Divida (D/V)", wacc_bloco.get("peso_divida"), None, None),
        (
            "w_capm",
            "WACC build-up CAPM",
            wacc_bloco.get("wacc_capm_buildup", wacc_bloco.get("wacc")),
            None,
            None,
        ),
    )
    for chave, rotulo, valor, _, _n in itens_wacc:
        formato = (
            FORMATO_MILHAR_2
            if chave in ("w_beta", "w_de", "w_beta_l")
            else FORMATO_PERCENTUAL_2
        )
        linha_valor(chave, rotulo, _numero(valor), formato=formato)

    def f_ke_brl() -> tuple[str, float | None]:
        ke_usd = _numero(wacc_bloco.get("ke_usd"))
        ipca = _numero(wacc_bloco.get("ipca"))
        cpi = _numero(wacc_bloco.get("cpi_eua"))
        formula = (
            f"=(1+{_ref('', 2, aba.L('w_ke_usd'))})*(1+{_ref('', 2, aba.L('w_ipca'))})"
            f"/(1+{_ref('', 2, aba.L('w_cpi'))})-1"
        )
        valor = (
            (1 + ke_usd) * (1 + ipca) / (1 + cpi) - 1
            if None not in (ke_usd, ipca, cpi)
            else None
        )
        return formula, valor

    formula_ke, ke_python = f_ke_brl()
    aba.calculo(
        aba.L("w_ke"),
        2,
        formula_ke,
        ke_python,
        _numero(wacc_bloco.get("ke_brl")),
        FORMATO_PERCENTUAL_2,
    )

    if manual:
        linha_valor(
            "w_manual",
            "WACC MANUAL do analista (usado)",
            wacc_usado,
            formato=FORMATO_PERCENTUAL_2,
            cor=COR_PREMISSA,
            negrito=True,
        )
        chave_wacc_usado = "w_manual"
    else:
        formula_wacc = (
            f"={_ref('', 2, aba.L('w_pe'))}*{_ref('', 2, aba.L('w_ke'))}"
            f"+{_ref('', 2, aba.L('w_pd'))}*{_ref('', 2, aba.L('w_kd'))}"
            f"*(1-{_ref('', 2, aba.L('w_t'))})"
        )
        pe, pd_, ke, kd, t_ = (
            _numero(wacc_bloco.get("peso_equity")),
            _numero(wacc_bloco.get("peso_divida")),
            _numero(wacc_bloco.get("ke_brl")),
            _numero(wacc_bloco.get("kd_historico")),
            _numero(wacc_bloco.get("aliquota_ir")),
        )
        wacc_python = (
            pe * ke + pd_ * kd * (1 - t_) if None not in (pe, pd_, ke, kd, t_) else None
        )
        linha_valor(
            "w_usado",
            "WACC USADO (VT/EV)",
            wacc_usado,
            formula=formula_wacc,
            valor_python=wacc_python,
            formato=FORMATO_PERCENTUAL_2,
            negrito=True,
        )
        chave_wacc_usado = "w_usado"

    g_valor = g
    linha_valor(
        "w_g",
        "g — perpetuidade",
        g_valor,
        formato=FORMATO_PERCENTUAL_2,
        cor=COR_PREMISSA,
    )

    # Nomes definidos usados nas formulas de VP/VT.
    wb.defined_names.add(
        DefinedName("WACC", attr_text=f"FCFF!$B${aba.L(chave_wacc_usado)}")
    )
    wb.defined_names.add(
        DefinedName("g_perpetuidade", attr_text=f"FCFF!$B${aba.L('w_g')}")
    )

    # ---- Retornos (multiplos + TIR/MOIC) ----
    cursor += 1
    aba.titulo(cursor, "RETORNOS DO ACIONISTA (multiplos, TIR, MOIC)", 10)
    cursor += 1
    multiplos = retornos.get("multiplos", {})
    aba.cabecalho(cursor, 2, [f"Ano {t}" for t in range(1, HORIZONTE_PROJECAO + 1)])
    cursor += 1
    for campo, rotulo in (
        ("ev_ebitda_preco_atual", "EV/EBITDA (preco atual)"),
        ("ev_ebitda_target", "EV/EBITDA (target)"),
        ("ev_receita_target", "EV/Receita (target)"),
        ("p_l_preco_atual", "P/L (preco atual)"),
    ):
        aba.rotulo(cursor, rotulo, recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            aba.numero(
                cursor,
                1 + t,
                _numero(multiplos.get(f"ano{t}", {}).get(campo)),
                FORMATO_MULTIPLO,
            )
        cursor += 1

    tir_moic = retornos.get("tir_moic", {})
    dividendos = tir_moic.get("dividendos_por_acao", {})
    horizonte_tir = int(tir_moic.get("horizonte_tir_anos", 5) or 5)
    cursor += 1
    aba.registrar("fluxo_acionista", cursor)
    aba.rotulo(
        cursor,
        "Fluxo do acionista (base): ano 0..N",
        recuo=1,
        comentario="-preco; +dividendos/acao; +preco de saida no ano N.",
    )
    fluxo = [-(_numero(tir_moic.get("preco_entrada")) or 0.0)]
    for t in range(1, horizonte_tir + 1):
        valor = _numero(dividendos.get(f"ano{t}")) or 0.0
        if t == horizonte_tir:
            valor += (
                _numero(tir_moic.get("cenarios", {}).get("base", {}).get("preco_saida"))
                or 0.0
            )
        fluxo.append(valor)
    for indice, valor in enumerate(fluxo):
        aba.numero(cursor, 2 + indice, valor, FORMATO_PRECO)
    linha_fluxo = cursor
    cursor += 1
    for nome in ("bear", "base", "bull"):
        cenario = tir_moic.get("cenarios", {}).get(nome, {})
        aba.rotulo(cursor, f"Cenario {nome}: saida | TIR | MOIC", recuo=1)
        aba.numero(cursor, 2, _numero(cenario.get("preco_saida")), FORMATO_PRECO)
        tir = _numero(cenario.get("tir_acionista"))
        if nome == "base" and tir is not None:
            faixa = (
                f"{_ref('', 2, linha_fluxo)}:"
                f"{_ref('', 2 + horizonte_tir, linha_fluxo)}"
            )
            aba.calculo(cursor, 3, f"=IRR({faixa})", tir, tir, FORMATO_PERCENTUAL)
        else:
            aba.numero(cursor, 3, tir, FORMATO_PERCENTUAL)
        aba.numero(cursor, 4, _numero(cenario.get("moic")), FORMATO_MULTIPLO)
        cursor += 1

    return aba


def _aba_fcfe(wb: Workbook, ctx: dict[str, Any], aba_modelo: Aba) -> Aba:
    """FCFE: LL + D&A - ΔWK - CAPEX + ΔDivida -> VP ao Ke -> Target (FCFE)."""
    aba = Aba(wb, "FCFE")
    ws = aba.ws
    ws.column_dimensions["A"].width = 40
    for coluna in range(2, 11):
        ws.column_dimensions[get_column_letter(coluna)].width = 14

    projecao = ctx["projecao"]
    fcfe = projecao["fcfe"]
    valuation = projecao.get("fcfe_valuation", {})
    ev_equity = projecao.get("ev_equity", {})
    ke = _numero(valuation.get("ke"))
    g = _numero(valuation.get("g"))
    conv = ctx["convencao_desconto"]

    def _exp(ano: int) -> float:
        """Expoente de desconto do ano (mesma convencao do motor)."""
        return expoente_desconto(ano, conv)

    def _exp_txt(ano: int) -> str:
        """Expoente formatado para a formula (8, nao 8.0, em periodo cheio)."""
        return f"{_exp(ano):g}"

    def fx(campo: str, t: int) -> float | None:
        return _numero(fcfe.get(f"ano{t}", {}).get(campo))

    aba.titulo(1, f"FCFE — DCF pelo acionista ({ctx['ticker']})", 10)
    aba.cabecalho(2, 2, [f"Ano {t}" for t in range(1, HORIZONTE_PROJECAO + 1)])
    cursor = 3

    def linha_serie(
        chave: str,
        rotulo: str,
        formula_fn: Callable[[int], tuple[str, float | None]] | None,
        valores_motor: Callable[[int], float | None],
        formato: str = FORMATO_MILHAR,
        negrito: bool = False,
    ) -> None:
        nonlocal cursor
        aba.registrar(chave, cursor)
        aba.rotulo(cursor, rotulo, negrito=negrito, recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            valor_motor = valores_motor(t)
            if formula_fn is None or valor_motor is None:
                aba.numero(cursor, 1 + t, valor_motor, formato, negrito=negrito)
                continue
            formula, valor_python = formula_fn(t)
            aba.calculo(
                cursor,
                1 + t,
                formula,
                valor_python,
                valor_motor,
                formato,
                negrito=negrito,
            )
        cursor += 1

    def rl(chave: str, t: int) -> str:
        return _ref("", 1 + t, aba.L(chave))

    linha_serie(
        "ll",
        "Lucro liquido (Modelo)",
        lambda t: (
            f"={_rm(ctx, aba_modelo, 'lucro_liquido', t)}",
            _proj(ctx, "dre", "lucro_liquido", t),
        ),
        lambda t: fx("lucro_liquido", t),
    )
    linha_serie(
        "da",
        "(+) Depreciacao e amortizacao",
        lambda t: (
            f"=-{_rm(ctx, aba_modelo, 'dre_da', t)}",
            fx("depreciacao_amortizacao", t),
        ),
        lambda t: fx("depreciacao_amortizacao", t),
    )
    linha_serie(
        "delta_wk",
        "(-) Δ Working Capital",
        lambda t: (
            f"=-{_rm(ctx, aba_modelo, 'wk_delta_nwc', t)}",
            -(fx("delta_nwc", t) or 0.0),
        ),
        lambda t: (
            -(fx("delta_nwc", t) or 0.0) if fx("delta_nwc", t) is not None else None
        ),
    )
    linha_serie(
        "capex",
        "(-) CAPEX",
        lambda t: (
            f"={_rm(ctx, aba_modelo, 'ppe_capex', t)}",
            _proj(ctx, "ppe", "capex", t),
        ),
        lambda t: (
            -(fx("capex_saida_caixa", t) or 0.0)
            if fx("capex_saida_caixa", t) is not None
            else None
        ),
    )
    linha_serie(
        "delta_divida",
        "(+) Δ Divida (captacoes - amortizacoes)",
        lambda t: (
            f"={_rm(ctx, aba_modelo, 'div_eop', t)}"
            f"-{_rm(ctx, aba_modelo, 'div_bop', t)}",
            (_proj(ctx, "divida", "divida_bruta", t) or 0.0)
            - (_proj(ctx, "divida", "divida_abertura", t) or 0.0),
        ),
        lambda t: fx("delta_divida", t),
    )

    def f_fcfe(t: int) -> tuple[str, float | None]:
        formula = (
            f"={rl('ll', t)}+{rl('da', t)}+{rl('delta_wk', t)}"
            f"+{rl('capex', t)}+{rl('delta_divida', t)}"
        )
        partes = [
            fx("lucro_liquido", t),
            fx("depreciacao_amortizacao", t),
            -(fx("delta_nwc", t) or 0.0),
            -(fx("capex_saida_caixa", t) or 0.0),
            fx("delta_divida", t),
        ]
        valor = (
            sum(x for x in partes if x is not None) if partes[0] is not None else None
        )
        return formula, valor

    linha_serie("fcfe", "(=) FCFE", f_fcfe, lambda t: fx("fcfe", t), negrito=True)
    linha_serie("t_anos", "t (desconto)", None, lambda t: _exp(t), FORMATO_MILHAR_2)

    def f_vp(t: int) -> tuple[str, float | None]:
        fluxo = fx("fcfe", t)
        valor = fluxo / (1 + ke) ** _exp(t) if fluxo is not None and ke else None
        return f"={rl('fcfe', t)}/(1+Ke_fcfe)^{rl('t_anos', t)}", valor

    linha_serie(
        "vp_fcfe",
        "VP do FCFE (ao Ke)",
        f_vp,
        lambda t: ((fx("fcfe", t) or 0.0) / (1 + ke) ** _exp(t) if ke else None),
    )

    cursor += 1
    aba.titulo(cursor, "EQUITY VALUE DIRETO -> TARGET (FCFE)", 10)
    cursor += 1

    def linha_valor(
        chave: str,
        rotulo: str,
        valor_motor: Any,
        formula: str | None = None,
        valor_python: Any = None,
        formato: str = FORMATO_MILHAR,
        cor: str = COR_FORMULA,
        negrito: bool = False,
    ) -> None:
        nonlocal cursor
        aba.registrar(chave, cursor)
        aba.rotulo(cursor, rotulo, negrito=negrito, recuo=1)
        if formula is not None:
            aba.calculo(
                cursor,
                2,
                formula,
                valor_python,
                valor_motor,
                formato,
                cor=cor,
                negrito=negrito,
            )
        else:
            aba.numero(cursor, 2, valor_motor, formato, cor=cor, negrito=negrito)
        cursor += 1

    linha_valor(
        "ke",
        "Ke BRL (CAPM Brasil, mesmo do build-up FCFF)",
        ke,
        formato=FORMATO_PERCENTUAL_2,
    )
    linha_valor(
        "g", "g — perpetuidade", g, formato=FORMATO_PERCENTUAL_2, cor=COR_PREMISSA
    )
    wb.defined_names.add(DefinedName("Ke_fcfe", attr_text=f"FCFE!$B${aba.L('ke')}"))

    soma_vp = _numero(valuation.get("soma_vp_fcfe"))
    soma_python = (
        sum(
            (fx("fcfe", t) or 0.0) / (1 + ke) ** _exp(t)
            for t in range(1, HORIZONTE_PROJECAO + 1)
        )
        if ke
        else None
    )
    linha_valor(
        "soma_vp",
        "PV do fluxo explicito (anos 1-8)",
        soma_vp,
        formula=f"=SUM({_ref('', 2, aba.L('vp_fcfe'))}:"
        f"{_ref('', 9, aba.L('vp_fcfe'))})",
        valor_python=soma_python,
    )

    # VT com ΔDivida NORMALIZADA: (FCFE8 - ΔD8 + g x D8) x (1+g) / (Ke - g).
    vt_bruto = _numero(valuation.get("vt_bruto"))
    fcfe8 = fx("fcfe", 8)
    delta_d8 = fx("delta_divida", 8)
    d8 = _proj(ctx, "divida", "divida_bruta", 8)
    base_python = (
        (fcfe8 - delta_d8 + g * d8) if None not in (fcfe8, delta_d8, g, d8) else None
    )
    vt_python = (
        base_python * (1 + g) / (ke - g)
        if base_python is not None and ke is not None and g is not None and ke > g
        else None
    )
    formula_vt = (
        f"=({rl('fcfe', 8)}-{rl('delta_divida', 8)}"
        f"+{_ref('', 2, aba.L('g'))}*{_rm(ctx, aba_modelo, 'div_eop', 8)})"
        f"*(1+{_ref('', 2, aba.L('g'))})/(Ke_fcfe-{_ref('', 2, aba.L('g'))})"
    )
    linha_valor(
        "vt",
        f"Valor terminal (base: {valuation.get('base_vt', 'n/d')})",
        vt_bruto,
        formula=formula_vt,
        valor_python=vt_python,
    )
    vp_vt = _numero(valuation.get("vp_vt"))
    linha_valor(
        "vp_vt",
        "PV da perpetuidade",
        vp_vt,
        formula=f"={_ref('', 2, aba.L('vt'))}/(1+Ke_fcfe)^{_exp_txt(8)}",
        valor_python=(
            vt_bruto / (1 + ke) ** _exp(8) if vt_bruto is not None and ke else None
        ),
    )
    equity_fcfe = _numero(valuation.get("equity_value_fcfe"))
    linha_valor(
        "equity",
        "(=) Equity Value (FCFE direto)",
        equity_fcfe,
        formula=f"={_ref('', 2, aba.L('soma_vp'))}+{_ref('', 2, aba.L('vp_vt'))}",
        valor_python=(
            (soma_vp or 0.0) + (vp_vt or 0.0)
            if soma_vp is not None and vp_vt is not None
            else None
        ),
        negrito=True,
    )

    acoes = _numero(ev_equity.get("acoes_fully_diluted"))
    fator = _numero(ev_equity.get("fator_escala_moeda")) or 1.0
    linha_valor("acoes", "Acoes fully diluted", acoes, formato="#,##0")
    linha_valor("fator", "Fator de escala", fator, formato="#,##0")
    target_fcfe = _numero(valuation.get("target_price_fcfe"))
    linha_valor(
        "target",
        "TARGET PRICE (FCFE)",
        target_fcfe,
        formula=f"={_ref('', 2, aba.L('equity'))}*{_ref('', 2, aba.L('fator'))}"
        f"/{_ref('', 2, aba.L('acoes'))}",
        valor_python=(
            equity_fcfe * fator / acoes if equity_fcfe is not None and acoes else None
        ),
        formato=FORMATO_PRECO,
        negrito=True,
    )
    preco = _numero(ev_equity.get("preco_atual"))
    linha_valor("preco", "Preco atual", preco, formato=FORMATO_PRECO)
    linha_valor(
        "upside",
        "Upside (FCFE)",
        (target_fcfe / preco - 1) if target_fcfe is not None and preco else None,
        formula=f"={_ref('', 2, aba.L('target'))}/{_ref('', 2, aba.L('preco'))}-1",
        valor_python=(
            target_fcfe / preco - 1 if target_fcfe is not None and preco else None
        ),
        formato=FORMATO_PERCENTUAL,
    )

    cursor += 1
    aba.titulo(cursor, "RECONCILIACAO FCFE x BRIDGE FCFF", 10)
    cursor += 1
    equity_bridge = _numero(valuation.get("equity_value_bridge"))
    linha_valor("eq_bridge", "Equity do bridge FCFF/WACC", equity_bridge)
    divergencia = _numero(valuation.get("divergencia_vs_bridge"))
    linha_valor(
        "divergencia",
        "Divergencia (FCFE / bridge - 1)",
        divergencia,
        formula=f"={_ref('', 2, aba.L('equity'))}/{_ref('', 2, aba.L('eq_bridge'))}-1",
        valor_python=(
            equity_fcfe / equity_bridge - 1
            if equity_fcfe is not None and equity_bridge
            else None
        ),
        formato=FORMATO_PERCENTUAL,
    )
    aba.texto(
        cursor,
        1,
        "Nota: FCFE/Ke e a CHECAGEM do bridge FCFF/WACC (metodo primario do "
        "target). Divergencia grande = estruturas de desconto diferentes "
        "(Ke unico vs WACC com pesos contabeis) — anotada, nao e erro.",
        cor=COR_CINZA_ND,
    )
    if valuation.get("aviso"):
        aba.texto(cursor + 1, 1, f"AVISO: {valuation['aviso']}", cor=COR_AMARELO_FUNDO)
    return aba


# ---------------------------------------------------------------------------
# Aba 6 — Macro
# ---------------------------------------------------------------------------


def _aba_macro(wb: Workbook, ctx: dict[str, Any]) -> Aba:
    """Bloco macro_anual + series coletadas (fonte e data)."""
    aba = Aba(wb, "Macro")
    ws = aba.ws
    ws.column_dimensions["A"].width = 34
    for coluna in range(2, 11):
        ws.column_dimensions[get_column_letter(coluna)].width = 12

    macro = ctx["macro"]
    macro_anual = macro.get("macro_anual", {})
    aba.titulo(1, "MACRO ANUAL (Focus + convergencia as metas)", 10)
    anos_cal = [
        macro_anual.get(f"ano{t}", {}).get("ano_calendario", f"Ano {t}")
        for t in range(1, HORIZONTE_PROJECAO + 1)
    ]
    aba.cabecalho(2, 2, [str(a) for a in anos_cal])
    linha = 3
    for campo, rotulo in (
        ("ipca", "IPCA esperado"),
        ("selic", "Selic esperada (fim de periodo)"),
        ("cdi", "CDI esperado (Selic - 0,1pp)"),
        ("pib", "PIB real esperado"),
    ):
        aba.rotulo(linha, rotulo, recuo=1)
        for t in range(1, HORIZONTE_PROJECAO + 1):
            aba.numero(
                linha,
                1 + t,
                _numero(macro_anual.get(f"ano{t}", {}).get(campo)),
                FORMATO_PERCENTUAL_2,
            )
        linha += 1
    aba.rotulo(linha, "Origem (por ano)", recuo=1)
    for t in range(1, HORIZONTE_PROJECAO + 1):
        aba.texto(
            linha,
            1 + t,
            str(macro_anual.get(f"ano{t}", {}).get("origem_selic", "n/d")),
            cor=COR_CINZA_ND,
        )
    linha += 2

    aba.titulo(linha, "SERIES COLETADAS (spot)", 10)
    linha += 1
    for campo, rotulo, formato in (
        ("selic_atual", "Selic meta vigente (SGS 432)", FORMATO_PERCENTUAL_2),
        ("cdi_atual", "CDI anualizado (SGS 4389)", FORMATO_PERCENTUAL_2),
        ("igpm_12m", "IGP-M acumulado 12m (SGS 189)", FORMATO_PERCENTUAL_2),
        ("cambio_brl_usd", "Cambio BRL/USD (SGS 1)", FORMATO_MILHAR_2),
    ):
        aba.rotulo(linha, rotulo, recuo=1)
        aba.numero(linha, 2, _numero(macro.get(campo)), formato)
        linha += 1
    aba.rotulo(linha, "Data da coleta", recuo=1)
    aba.texto(linha, 2, str(macro.get("data_coleta", "n/d")), cor=COR_CINZA_ND)
    linha += 1
    aba.rotulo(linha, "Fonte", recuo=1)
    aba.texto(linha, 2, "BACEN (SGS + Focus/Expectativas)", cor=COR_CINZA_ND)
    return aba


# ---------------------------------------------------------------------------
# Aba 7 — Sensibilidades
# ---------------------------------------------------------------------------


def _aba_sensibilidades(wb: Workbook, ctx: dict[str, Any]) -> Aba:
    """Sensibilidade Target x (WACC, g) + grade Bear/Base/Bull dos retornos."""
    aba = Aba(wb, "Sensibilidades")
    ws = aba.ws
    ws.column_dimensions["A"].width = 30

    projecao = ctx["projecao"]
    fcff = projecao["fcff"]
    ev_equity = projecao.get("ev_equity", {})
    vt = projecao.get("valor_terminal", {})
    wacc_base = _numero(projecao.get("wacc", {}).get("wacc")) or 0.12
    g_base = _numero(vt.get("g")) or 0.03
    ajustes = ev_equity.get("ajustes_bridge", {})
    acoes = _numero(ev_equity.get("acoes_fully_diluted")) or 1.0
    fator = _numero(ev_equity.get("fator_escala_moeda")) or 1.0
    fluxos = [
        _numero(fcff.get(f"ano{t}", {}).get("fcff")) or 0.0
        for t in range(1, HORIZONTE_PROJECAO + 1)
    ]
    ajuste_liquido = (
        -(_numero(ajustes.get("divida_bruta")) or 0.0)
        + (_numero(ajustes.get("caixa_equivalentes")) or 0.0)
        + (_numero(ajustes.get("aplicacoes_financeiras")) or 0.0)
        - (_numero(ajustes.get("participacoes_minoritarias")) or 0.0)
        + (_numero(ajustes.get("investimentos_coligadas")) or 0.0)
        + (_numero(ajustes.get("ativos_nao_operacionais")) or 0.0)
    )

    conv = ctx["convencao_desconto"]

    def target_para(wacc: float, g: float) -> float | None:
        """Recalculo de APRESENTACAO da grade (mesma matematica do motor)."""
        if wacc <= g or wacc <= 0:
            return None
        soma_vp = sum(
            f / (1 + wacc) ** expoente_desconto(t, conv)
            for t, f in enumerate(fluxos, start=1)
        )
        base = fluxos[-1]
        if base < 0:
            base = _numero(fcff.get("ano8", {}).get("nopat")) or 0.0
        vt_bruto = base * (1 + g) / (wacc - g)
        ev = soma_vp + vt_bruto / (1 + wacc) ** expoente_desconto(
            HORIZONTE_PROJECAO, conv
        )
        return (ev + ajuste_liquido) * fator / acoes

    passos_wacc = [-0.015, -0.01, -0.005, 0.0, 0.005, 0.01, 0.015]
    passos_g = [-0.01, -0.005, 0.0, 0.005, 0.01]
    aba.titulo(1, "SENSIBILIDADE — TARGET x (WACC, g)", 9)
    aba.cabecalho(2, 2, [f"g {(g_base + dg) * 100:.1f}%" for dg in passos_g])
    linha = 3
    preco = _numero(ev_equity.get("preco_atual")) or 0.0
    primeira = (3, 2)
    ultima = (3 + len(passos_wacc) - 1, 2 + len(passos_g) - 1)
    for dw in passos_wacc:
        aba.rotulo(linha, f"WACC {(wacc_base + dw) * 100:.2f}%")
        for indice, dg in enumerate(passos_g):
            alvo = target_para(wacc_base + dw, g_base + dg)
            aba.numero(linha, 2 + indice, alvo, FORMATO_PRECO)
        linha += 1

    if preco > 0:
        faixa = (
            f"{get_column_letter(primeira[1])}{primeira[0]}:"
            f"{get_column_letter(ultima[1])}{ultima[0]}"
        )
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(
                operator="greaterThan",
                formula=[str(preco * (1 + LIMITE_COMPRA))],
                fill=PatternFill("solid", start_color=COR_VERDE_FUNDO),
            ),
        )
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(
                operator="lessThan",
                formula=[str(preco * (1 + LIMITE_VENDA))],
                fill=PatternFill("solid", start_color=COR_VERMELHO_FUNDO),
            ),
        )
        ws.conditional_formatting.add(
            faixa,
            CellIsRule(
                operator="between",
                formula=[
                    str(preco * (1 + LIMITE_VENDA)),
                    str(preco * (1 + LIMITE_COMPRA)),
                ],
                fill=PatternFill("solid", start_color=COR_AMARELO_FUNDO),
            ),
        )

    linha += 1
    aba.titulo(linha, "GRADE BEAR / BASE / BULL (saida no ano 5)", 9)
    linha += 1
    aba.cabecalho(linha, 2, ["Preco de saida", "TIR", "MOIC"])
    linha += 1
    tir_moic = projecao.get("retornos", {}).get("tir_moic", {})
    for nome in ("bear", "base", "bull"):
        cenario = tir_moic.get("cenarios", {}).get(nome, {})
        aba.rotulo(linha, nome.capitalize(), negrito=(nome == "base"))
        aba.numero(linha, 2, _numero(cenario.get("preco_saida")), FORMATO_PRECO)
        aba.numero(linha, 3, _numero(cenario.get("tir_acionista")), FORMATO_PERCENTUAL)
        aba.numero(linha, 4, _numero(cenario.get("moic")), FORMATO_MULTIPLO)
        linha += 1
    aba.texto(
        linha + 1,
        1,
        "Grade de sensibilidade (apresentacao): mesma matematica do motor "
        "sobre os fluxos persistidos; o caso-base bate com a aba FCFF.",
        cor=COR_CINZA_ND,
    )
    return aba


# ---------------------------------------------------------------------------
# Aba 8 — Avisos
# ---------------------------------------------------------------------------


def _aba_avisos(wb: Workbook, ctx: dict[str, Any]) -> Aba:
    """Qualidade dos dados, checklist do motor, auditoria CVM e fallbacks."""
    aba = Aba(wb, "Avisos")
    ws = aba.ws
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30

    meta = ctx["metadados"]
    projecao = ctx["projecao"]
    aba.titulo(1, "QUALIDADE DOS DADOS E VERIFICACOES", 6)
    linha = 3
    aba.rotulo(linha, "Score", negrito=True)
    aba.texto(
        linha,
        2,
        f"{meta.get('score_confiabilidade', 'n/d')}/100 "
        f"(fonte: coleta CVM + relatorio de qualidade)",
    )
    linha += 1
    if ctx["premissas"].get("premissas_automaticas"):
        aba.rotulo(linha, "ALERTA", negrito=True)
        aba.texto(
            linha,
            2,
            "PREMISSAS AUTOMATICAS DE PARTIDA em uso — revise antes de "
            "usar como tese.",
            cor=COR_VERMELHO_FUNDO,
            negrito=True,
        )
        linha += 1
    linha += 1

    aba.titulo(linha, "CHECKLIST DO MOTOR", 6)
    linha += 1
    aba.cabecalho(linha, 1, ["ID", "Descricao", "Status", "Valor"])
    linha += 1
    cores_status = {
        "OK": COR_VERDE_FUNDO,
        "ERRO": COR_VERMELHO_FUNDO,
        "AVISO": COR_AMARELO_FUNDO,
        "ALERTA": COR_AMARELO_FUNDO,
    }
    for item in projecao.get("checklist", {}).get("itens", []):
        aba.texto(linha, 1, str(item.get("id", "")))
        aba.texto(linha, 2, str(item.get("descricao", "")))
        status = str(item.get("status", ""))
        celula = ws.cell(row=linha, column=3, value=status)
        celula.font = Font(name=FONTE_TEXTO, size=10, bold=True, color=COR_BRANCO)
        celula.fill = PatternFill(
            "solid", start_color=cores_status.get(status, COR_AZUL_ANCORA)
        )
        aba.valores[(linha, 3)] = status
        valor = item.get("valor")
        aba.texto(
            linha,
            4,
            (
                f"{valor:,.4g}"
                if isinstance(valor, (int, float)) and not isinstance(valor, bool)
                else str(valor)
            ),
        )
        linha += 1
    linha += 1

    auditoria = ctx["auditoria"]
    aba.titulo(linha, "AUDITORIA CVM (9.0.1)", 6)
    linha += 1
    if auditoria:
        contagem = auditoria.get("contagem", {})
        aba.rotulo(linha, "Status", negrito=True)
        aba.texto(
            linha,
            2,
            f"{auditoria.get('status_geral', 'n/d')} — "
            f"OK {contagem.get('OK', 0)} | AVISO {contagem.get('AVISO', 0)}"
            f" | ERRO {contagem.get('ERRO', 0)}",
        )
        linha += 1
        for item in auditoria.get("itens", []):
            if item.get("status") == "OK":
                continue
            aba.texto(
                linha,
                1,
                str(item.get("status", "")),
                cor=(
                    COR_AMARELO_FUNDO
                    if item.get("status") == "AVISO"
                    else COR_VERMELHO_FUNDO
                ),
                negrito=True,
            )
            aba.texto(
                linha, 2, f"{item.get('exercicio', '')}: {item.get('descricao', '')}"
            )
            linha += 1
    else:
        aba.texto(
            linha, 1, "Auditoria nao executada para este ticker.", cor=COR_CINZA_ND
        )
        linha += 1
    linha += 1

    aba.titulo(linha, "POLITICAS E FALLBACKS DO MOTOR", 6)
    linha += 1
    politicas = projecao.get("politicas_projecao", {})
    divida_pol = politicas.get("divida_balanco", {})
    wacc_bloco = projecao.get("wacc", {})
    fallbacks = (
        ("Modo da DRE", str(projecao.get("modo_dre", "n/d"))),
        (
            "Modo do capital de giro",
            str(projecao.get("wk", {}).get("ano1", {}).get("modo_capital_giro", "n/d")),
        ),
        (
            "Origem da taxa de aplicacao",
            str(divida_pol.get("origem_taxa_aplicacao", "n/d")),
        ),
        (
            "Origem do custo da divida",
            str(divida_pol.get("origem_custo_divida", "n/d")),
        ),
        ("Origem do payout", str(divida_pol.get("origem_payout", "n/d"))),
        ("Origem do WACC", str(wacc_bloco.get("wacc_origem", "build_up_capm"))),
        ("Instrumentos de divida", str(divida_pol.get("instrumentos_divida", "n/d"))),
    )
    for rotulo, valor in fallbacks:
        aba.rotulo(linha, "", negrito=False)
        aba.texto(linha, 1, rotulo)
        aba.texto(linha, 2, valor, cor=COR_CINZA_ND)
        linha += 1
    return aba


# ---------------------------------------------------------------------------
# Preview para o app (mesma fonte de dados; valores, sem formulas)
# ---------------------------------------------------------------------------


def _preview_de_aba(aba: Aba) -> list[tuple[str, pd.DataFrame]]:
    """Converte a matriz de valores rastreada em DataFrame unico da aba."""
    if not aba.valores:
        return [("(vazia)", pd.DataFrame())]
    max_linha = max(linha for linha, _ in aba.valores)
    max_coluna = max(coluna for _, coluna in aba.valores)
    grade = []
    for linha in range(1, max_linha + 1):
        registro = {}
        vazia = True
        for coluna in range(1, max_coluna + 1):
            valor = aba.valores.get((linha, coluna))
            if valor is not None:
                vazia = False
            registro[get_column_letter(coluna)] = valor
        if not vazia:
            grade.append(registro)
    quadro = pd.DataFrame(grade).fillna("")
    quadro = quadro.astype(str).replace("None", "")
    return [("Conteudo da aba", quadro)]


def montar_preview_por_aba(
    ticker: str,
    raiz_projeto: Path | None = None,
) -> dict[str, list[tuple[str, pd.DataFrame]]]:
    """Preview das 8 abas como DataFrames de VALORES para o app.

    Gera o workbook em memoria (mesma construcao do .xlsx) e converte a
    matriz de valores rastreada de cada aba — preview e arquivo nunca
    divergem.
    """
    _, abas = _construir_workbook(ticker, raiz_projeto)
    return {nome: _preview_de_aba(aba) for nome, aba in abas.items()}


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------


def _construir_workbook(
    ticker: str,
    raiz_projeto: Path | None = None,
) -> tuple[Workbook, dict[str, Aba]]:
    """Monta o workbook completo e devolve (wb, abas construidas)."""
    ctx = montar_contexto(ticker, raiz_projeto)
    wb = Workbook()
    wb.remove(wb.active)

    abas: dict[str, Aba] = {}
    abas["Capa"] = _aba_capa(wb, ctx)
    aba_premissas = _aba_premissas(wb, ctx)
    abas["Premissas"] = aba_premissas
    ctx["_aba_premissas"] = aba_premissas
    aba_modelo = _aba_modelo(wb, ctx)
    abas["Modelo"] = aba_modelo
    abas["FCFF"] = _aba_fcff(wb, ctx, aba_modelo)
    abas["FCFE"] = _aba_fcfe(wb, ctx, aba_modelo)
    abas["Macro"] = _aba_macro(wb, ctx)
    abas["Sensibilidades"] = _aba_sensibilidades(wb, ctx)
    abas["Avisos"] = _aba_avisos(wb, ctx)
    return wb, abas


def exportar_excel(ticker: str, raiz_projeto: Path | None = None) -> Path:
    """Gera o Excel 'Modelo' de 8 abas (Prompt 9.0.5) para o ticker.

    Le exclusivamente os resultados persistidos pelo motor; devolve o
    caminho do arquivo salvo em ``outputs/excel/<TICKER>_dcf.xlsx``.
    """
    raiz = resolver_raiz(raiz_projeto)
    ticker_normalizado = normalizar_ticker(ticker)
    wb, _ = _construir_workbook(ticker_normalizado, raiz)
    caminho = caminho_excel(ticker_normalizado, raiz)
    wb.save(caminho)
    logger.info("Excel gerado: %s", caminho)
    return caminho


def main() -> None:
    """Exporta o Excel padrao para DIRR3."""
    print(exportar_excel("DIRR3"))


if __name__ == "__main__":
    main()
