"""Testes do exportador Excel de 7 abas (secao 5.10 do ROTEIRO).

Usa os dados reais persistidos em data/processed (pipeline da Semana 3 ja
executado), como test_app.py. Valida o contrato estrutural do arquivo:
7 abas na ordem certa, formulas nativas nas abas de calculo, convencao de
cores WSP na aba de premissas, nomes definidos (WACC e g) e formatacao
condicional nas sensibilidades. A igualdade numerica formula x motor e
garantida em construcao por ``escrever_calculo`` (formula so entra quando
reproduz o valor persistido).
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.exportacao.exportador_excel import (
    COR_FONTE_INPUT,
    LINHA_MODELO_EBIT,
    LINHA_MODELO_RECEITA,
    LINHA_PPE_DA,
    LINHA_DIV_RF,
    exportar_excel,
)

RAIZ_PROJETO = Path(__file__).resolve().parents[1]

PROJECAO_DISPONIVEL = (
    RAIZ_PROJETO / "data" / "processed" / "DIRR3_projecao.json"
).exists()

pytestmark = pytest.mark.skipif(
    not PROJECAO_DISPONIVEL,
    reason="Pipeline da Semana 3 nao executado (data/processed ausente).",
)

ABAS_ESPERADAS = [
    "Capa",
    "Premissas",
    "Modelo Integrado",
    "Schedules",
    "Valuation",
    "Sensibilidades",
    "Output",
]


@pytest.fixture(scope="module")
def caminho_excel_dirr3() -> Path:
    """Gera (uma vez por modulo) o Excel de DIRR3 e devolve o caminho."""
    return exportar_excel("DIRR3", RAIZ_PROJETO)


@pytest.fixture(scope="module")
def workbook_dirr3(caminho_excel_dirr3: Path):
    """Workbook aberto preservando as formulas (data_only=False)."""
    return load_workbook(caminho_excel_dirr3, data_only=False)


def test_sete_abas_na_ordem(workbook_dirr3) -> None:
    """O arquivo tem exatamente as 7 abas da secao 5.10, na ordem."""
    assert workbook_dirr3.sheetnames == ABAS_ESPERADAS


def test_layout_fixo_alinhado_com_constantes(workbook_dirr3) -> None:
    """Os rotulos conferem com as constantes LINHA_* usadas nas formulas.

    Se o layout mudar sem atualizar as constantes, as formulas entre abas
    apontariam para celulas erradas — este teste trava o contrato.
    """
    modelo = workbook_dirr3["Modelo Integrado"]
    schedules = workbook_dirr3["Schedules"]
    assert modelo.cell(row=LINHA_MODELO_RECEITA, column=1).value == "Receita liquida"
    assert modelo.cell(row=LINHA_MODELO_EBIT, column=1).value == "EBIT"
    assert schedules.cell(row=LINHA_PPE_DA, column=1).value == "D&A do periodo"
    assert "Resultado financeiro" in str(
        schedules.cell(row=LINHA_DIV_RF, column=1).value
    )


def test_modelo_integrado_tem_formulas_nativas(workbook_dirr3) -> None:
    """Receita projetada usa formula ligada a aba Premissas (padrao WSP)."""
    modelo = workbook_dirr3["Modelo Integrado"]
    celula_ano1 = modelo.cell(row=LINHA_MODELO_RECEITA, column=5)
    assert str(celula_ano1.value).startswith("=")
    assert "Premissas!" in str(celula_ano1.value)


def test_premissas_com_fonte_azul_de_input(workbook_dirr3) -> None:
    """Os 8 valores individuais de crescimento usam fonte azul (input)."""
    premissas = workbook_dirr3["Premissas"]
    linha_crescimento = 5
    for coluna in range(3, 11):
        celula = premissas.cell(row=linha_crescimento, column=coluna)
        assert isinstance(celula.value, float)
        assert celula.font.color.rgb == COR_FONTE_INPUT


def test_nomes_definidos_wacc_e_g(workbook_dirr3) -> None:
    """WACC e g existem como nomes definidos (base para Data Tables)."""
    nomes = set(workbook_dirr3.defined_names)
    assert "WACC" in nomes
    assert "g_perpetuidade" in nomes


def test_sensibilidades_com_formatacao_condicional(workbook_dirr3) -> None:
    """As 3 tabelas geram regras condicionais (verde/amarelo/vermelho)."""
    sensibilidades = workbook_dirr3["Sensibilidades"]
    total_regras = sum(
        len(regras.rules) for regras in sensibilidades.conditional_formatting
    )
    assert total_regras >= 9  # 3 tabelas x 3 faixas de recomendacao


def test_valuation_e_output_com_imagens(workbook_dirr3) -> None:
    """Football Field/Waterfall e Dashboard/ROIC embutidos como PNG."""
    assert len(workbook_dirr3["Valuation"]._images) == 2
    assert len(workbook_dirr3["Output"]._images) == 2


def test_output_traz_checklist_com_status(workbook_dirr3) -> None:
    """A aba Output contem a tabela do checklist com IDs U1..NF5."""
    output = workbook_dirr3["Output"]
    ids = {
        str(output.cell(row=linha, column=1).value)
        for linha in range(1, output.max_row + 1)
    }
    assert "U1" in ids
    assert "NF1" in ids
